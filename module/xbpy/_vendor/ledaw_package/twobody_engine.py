import os
import re
import openpyxl
import numpy as np
import pandas as pd
from .classes import *
from .nbody_engine import extract_coords_from_line, normalize_path, fragments_equal, extract_final_fragment_labels_from_summary
from .nbody_to_twobody_engine import extract_coordinates_from_onebody_file


patterns = Patterns()


def get_two_body_filenames(two_body_orcaout_directory):
    """Reads all files from the specified directory and saves their full paths to a list."""
    # Normalize the directory path to remove any trailing slashes
    two_body_orcaout_directory = os.path.normpath(two_body_orcaout_directory)

    # Initialize an empty list to store file paths
    two_body_orcaout_filenames = []

    # Walk through the directory and collect file paths
    for root, _, files in os.walk(two_body_orcaout_directory):
        for file in files:
            # Combine the root path with the file name to get the full file path
            full_path = os.path.normpath(os.path.join(root, file))
            two_body_orcaout_filenames.append(full_path)

    return two_body_orcaout_filenames


def extract_coords_from_one_body_files(one_body_orcaout_filenames, tolerance=1e-3):
    """Extract coordinates from one-body files based on first non-ghost line after *xyz 
    and in its absence after CARTESIAN COORDINATES (ANGSTROEM). Returns filename → coords."""
    label_mapping = {}

    for filename in one_body_orcaout_filenames:
        coords = None
        with open(filename, 'r') as file:
            content = file.readlines()

        # Try *xyz block first
        found_xyz = False
        for i, line in enumerate(content):
            if re.search(r"\*\s*xyz\b", line, re.IGNORECASE):
                found_xyz = True
                for subsequent_line in content[i + 1:]:
                    if ':' not in subsequent_line:
                        numbers = re.findall(r'-?\d+\.\d+', subsequent_line)
                        if len(numbers) >= 3:
                            coords = [float(num) for num in numbers[:3]]
                        break
                break

        # if *xyz block not found, try the "CARTESIAN COORDINATES (ANGSTROEM)" block
        if not found_xyz:
            in_cartesian = False
            for line in content:
                if "CARTESIAN COORDINATES (ANGSTROEM)" in line:
                    in_cartesian = True
                    continue
                if in_cartesian and (re.match(r"-{5,}", line) or line.strip() == ""):
                    continue
                if in_cartesian:
                    parts = line.strip().split()
                    if len(parts) >= 4:
                        try:
                            coords = [float(val) for val in parts[1:4]]
                            break  # Only first line
                        except ValueError:
                            continue
                    elif line.strip() == "":
                        break

        if coords:
            label_mapping[filename] = coords

    return label_mapping


def extract_labels_from_two_body_files(two_body_orcaout_directory, label_mapping, tolerance=1e-3):
    """Extract labels from two-body files using coordinate comparison with a specified tolerance.
    Stops searching once 'INTERNAL COORDINATES (ANGSTROEM)' is encountered and looks for max two FRAGMENT X with coordinates.
    Match fragment coords in two-body files to one-body coords and assign correct labels using filename_to_label."""
    
    filename_to_label = {fname: idx for idx, fname in enumerate(label_mapping.keys(), start=1)}
    two_body_labels = {}

    for file_name in os.listdir(two_body_orcaout_directory):
        file_path = os.path.join(two_body_orcaout_directory, file_name)
        if not os.path.isfile(file_path):
            continue
        
        with open(file_path, 'r') as file:
            lines = file.readlines()

        labels = []
        fragments_with_coords = 0
        fragment_started = False

        for line in lines:
            line = line.strip()

            if "INTERNAL COORDINATES (ANGSTROEM)" in line:
                break

            if line.startswith("FRAGMENT"):
                fragment_started = True

            if fragment_started:
                parts = line.split()
                if len(parts) == 4:
                    try:
                        coords = list(map(float, parts[1:4]))
                    except ValueError:
                        continue

                    for fname, known_coords in label_mapping.items():
                        if all(abs(a - b) <= tolerance for a, b in zip(coords, known_coords)):
                            labels.append(filename_to_label[fname])
                            break

                    fragments_with_coords += 1

            if fragments_with_coords == 2 and len(labels) == 2:
                break

        if len(labels) == 2:
            two_body_labels[file_name] = labels

    return two_body_labels


def detect_bsse_from_onebody_files(one_body_orcaout_filenames):
    """Determine if BSSE is applied by checking the first one-body file for a line containing ':' in *xyz block (ghost atoms).
    Returns True if BSSE is detected, else False."""
    if not one_body_orcaout_filenames:
        return False

    first_file = one_body_orcaout_filenames[0]
    in_xyz_block = False

    with open(first_file, 'r') as file:
        for line in file:
            if '*xyz' in line.lower():
                in_xyz_block = True
                continue

            if in_xyz_block:
                if line.strip().startswith('*') or "END OF INPUT" in line:
                    break

                if ':' in line:
                    return True

    return False


def extract_real_and_ghost_coords(filepath):
    real_coords = []
    ghost_coords = []
    recording = False

    with open(filepath, 'r') as file:
        for line in file:
            if re.search(r"\*\s*xyz\b", line, re.IGNORECASE):
                recording = True
                continue
            elif recording and (line.strip().startswith('*') or "END OF INPUT" in line):
                break
            elif recording:
                coords = extract_coords_from_line(line)
                if coords is None:
                    continue

                parts = line.strip().split()
                if not parts:
                    continue

                symbol = next((p for p in parts if re.match(r"[A-Z][a-z]?:?", p)), "")

                if ':' in symbol:
                    ghost_coords.append(tuple(round(x, 6) for x in coords))
                else:
                    real_coords.append(tuple(round(x, 6) for x in coords))

    return real_coords, ghost_coords


def extract_real_coords_from_twobody_file(filepath):
    with open(filepath, 'r') as file:
        lines = file.readlines()

    fragments = []
    current_fragment = []
    recording = False

    for line in lines:
        if "INTERNAL COORDINATES (ANGSTROEM)" in line:
            break  # Stop parsing before internal coord section

        line = line.strip()

        if line.startswith("FRAGMENT"):
            if current_fragment:
                fragments.append(current_fragment)
                current_fragment = []
            recording = True
            continue

        if recording:
            parts = line.split()
            if len(parts) == 4:
                try:
                    xyz = [float(x) for x in parts[1:4]]
                    current_fragment.append(tuple(round(val, 6) for val in xyz))
                except ValueError:
                    continue

    if current_fragment:
        fragments.append(current_fragment)

    if len(fragments) != 2:
        raise ValueError(f"Expected 2 fragments in file {filepath}, but found {len(fragments)}")

    return fragments[0], fragments[1]


def generate_pairwise_fragment_index_map(one_body_orcaout_filenames, two_body_orcaout_directory, tolerance=1e-3):
    """Generates a mapping of two-body output filenames to fragment index pairs by matching coordinates from one-body files,
    and returns the mapping, index assignments, and ordered one-body filenames."""

    filename_to_coords = {}
    for filename in one_body_orcaout_filenames:
        coords = extract_coordinates_from_onebody_file(filename)
        if coords:
            coords = [tuple(float(f"{c:.6f}") for c in triplet) for triplet in coords]
            filename_to_coords[filename] = coords

    unique_fragments = []
    coords_to_index = {}
    filename_to_index = {}
    current_index = 1

    # Assign a unique index per unique coordinate
    for filename, coords in filename_to_coords.items():
        found = False
        for known_coords in unique_fragments:
            if fragments_equal(known_coords, coords, tol=tolerance):
                assigned_index = coords_to_index[tuple(map(tuple, known_coords))]
                filename_to_index[filename] = assigned_index
                found = True
                break
        if not found:
            unique_fragments.append(coords)
            coords_to_index[tuple(map(tuple, coords))] = current_index
            filename_to_index[filename] = current_index
            current_index += 1

    # Match pairs from two-body files
    two_body_labels = {}
    used_pairs = set()

    for fname in os.listdir(two_body_orcaout_directory):
        file_path = os.path.join(two_body_orcaout_directory, fname)
        if not os.path.isfile(file_path):
            continue

        with open(file_path, 'r') as f:
            lines = f.readlines()

        fragment_coords_list = []
        current_fragment = []
        in_fragment = False

        for line in lines:
            if "INTERNAL COORDINATES (ANGSTROEM)" in line:
                break

            line = line.strip()
            if line.startswith("FRAGMENT"):
                if current_fragment:
                    fragment_coords_list.append(current_fragment)
                current_fragment = []
                in_fragment = True

            elif in_fragment and len(line.split()) == 4:
                try:
                    xyz = [float(x) for x in line.split()[1:4]]
                    current_fragment.append(tuple(round(v, 6) for v in xyz))
                except:
                    continue

        if current_fragment:
            fragment_coords_list.append(current_fragment)

        if len(fragment_coords_list) != 2:
            continue

        frag_indices = []
        for coords in fragment_coords_list:
            matched_index = None
            for known_coords in unique_fragments:
                if fragments_equal(coords, known_coords, tol=tolerance):
                    matched_index = coords_to_index[tuple(map(tuple, known_coords))]
                    break
            if matched_index is not None:
                frag_indices.append(matched_index)

        if len(frag_indices) == 2:
            pair = tuple(sorted(frag_indices))
            if pair not in used_pairs:
                two_body_labels[fname] = frag_indices
                used_pairs.add(pair)

    index_filename_pairs = [(idx, fname) for fname, idx in filename_to_index.items()]
    index_filename_pairs.sort()
    ordered_filenames = [fname for _, fname in index_filename_pairs]

    return two_body_labels, filename_to_index, ordered_filenames


def generate_bsse_onebody_file_pair(two_body_files, one_body_files, fragment_index_map, extract_real_coords_from_twobody_file, extract_real_and_ghost_coords, fragments_equal, tol=1e-3):
    """For each two-body file (with two real fragments), find the corresponding one-body file pair using coordinate matching."""

    onebody_data = []
    for one_body_file in one_body_files:
        real_coords, ghost_coords = extract_real_and_ghost_coords(one_body_file)
        onebody_data.append((one_body_file, real_coords, ghost_coords))

    bsse_pair_map = {}

    for two_body_file in two_body_files:
        try:
            fragA_coords, fragB_coords = extract_real_coords_from_twobody_file(two_body_file)
        except Exception as e:
            print(f"Warning: Skipping {two_body_file} — {e}")
            continue

        match_A_real = None
        match_B_real = None

        for one_body_file, real_coords, ghost_coords in onebody_data:
            if fragments_equal(real_coords, fragA_coords, tol) and fragments_equal(ghost_coords, fragB_coords, tol):
                match_A_real = one_body_file
            elif fragments_equal(real_coords, fragB_coords, tol) and fragments_equal(ghost_coords, fragA_coords, tol):
                match_B_real = one_body_file

        if match_A_real and match_B_real:
            index_A = fragment_index_map.get(match_A_real)
            index_B = fragment_index_map.get(match_B_real)

            if index_A is None or index_B is None:
                print(f"Warning: Fragment indices missing for {match_A_real} or {match_B_real}")
                continue

            bsse_pair_map[(index_A, index_B)] = (match_A_real, match_B_real)
        else:
            print(f"Warning: No complete one-body match for {two_body_file}")

    return bsse_pair_map


def extract_onebody_dielectric_or_cds_values(one_body_orcaout_filenames, pattern):
    """Extract dielectric or CDS values from one-body files indexed by file order (1-based)."""
    values = {}
    for idx, filename in enumerate(one_body_orcaout_filenames, start=1):
        try:
            with open(normalize_path(filename), 'r') as file:
                content = file.read()
                match = re.search(pattern, content)
                if match:
                    values[idx] = float(match.group(1))
                else:
                    values[idx] = None # Explicitly set to None if not found
        except FileNotFoundError:
            print(f"One-body file not found: {filename}")
            values[idx] = None
        except Exception as e:
            print(f"Error processing one-body file {filename}: {e}")
            values[idx] = None
    return values


def extract_twobody_dielectric_or_cds_values(two_body_orcaout_directory, two_body_labels, pattern):
    """Extract dielectric or CDS values from two-body files."""
    values = {}
    for filename, (label1, label2) in two_body_labels.items():
        file_path = os.path.join(two_body_orcaout_directory, filename)
        if not os.path.isfile(file_path):
            print(f"Two-body file not found: {file_path}")
            values[filename] = (label1, label2, None) # Explicitly set to None if not found
            continue
        try:
            with open(file_path, 'r') as file:
                content = file.read()
                match = re.search(pattern, content)
                if match:
                    values[filename] = (label1, label2, float(match.group(1)))
                else:
                    values[filename] = (label1, label2, None) # Explicitly set to None if not found
        except Exception as e:
            print(f"Error processing two-body file {file_path}: {e}")
            values[filename] = (label1, label2, None)
    return values


def populate_twobody_ref_diel_matrices_non_bsse(one_body_orcaout_filenames, two_body_orcaout_directory, conversion_factor, two_body_labels, LEDAW_output_path_two_body):
    """Populate the two-body REF DIEL interaction matrix and save it to an Excel file. Return the matrix and its sum."""

    normalized_LEDAW_output_path_two_body = normalize_path(LEDAW_output_path_two_body)

    dielectric_pattern = r"CPCM Dielectric\s*:\s*([-+]?\d*\.\d+)\s*Eh"

    one_body_ref_diel_values = extract_onebody_dielectric_or_cds_values(one_body_orcaout_filenames, dielectric_pattern)
    two_body_diel_values = extract_twobody_dielectric_or_cds_values(two_body_orcaout_directory, two_body_labels, dielectric_pattern)

    n = len(one_body_orcaout_filenames)
    ref_diel_matrix = np.zeros((n, n))

    all_diel_found = True
    for idx in range(1, n + 1):
        if one_body_ref_diel_values.get(idx) is None:
            all_diel_found = False
    for filename in two_body_labels.keys():
        if two_body_diel_values.get(filename, (None, None, None))[2] is None:
            all_diel_found = False

    if not all_diel_found:
        print("Warning: At least one Dielectric component was not found. Hence, the entire CPCM contribution matrix will be set to zero.")
        ref_diel_matrix[:] = 0
        ref_diel_int_energy = np.nansum(ref_diel_matrix)
        if ref_diel_int_energy != 0:
            df = pd.DataFrame(ref_diel_matrix, index=range(1, n + 1), columns=range(1, n + 1))
            output_file_path = os.path.join(normalized_LEDAW_output_path_two_body, 'SOLV-fp.xlsx')
            normalized_output_file_path = normalize_path(output_file_path)
            os.makedirs(os.path.dirname(normalized_output_file_path), exist_ok=True)
            
            writer_args = {'engine': 'openpyxl'}
            if os.path.exists(normalized_output_file_path):
                writer_args['mode'] = 'a'
                writer_args['if_sheet_exists'] = 'replace'
            else:
                writer_args['mode'] = 'w'

            with pd.ExcelWriter(normalized_output_file_path, **writer_args) as writer:
                df.to_excel(writer, sheet_name='REF-DIEL')
            print(f"Two-body REF dielectric LED interaction energy matrix was written to {normalized_output_file_path}")
        return ref_diel_matrix, ref_diel_int_energy

    for filename, (label1, label2, two_body_diel_value) in two_body_diel_values.items():
        two_body_diel_value_to_use = two_body_diel_value or 0.0
        one_body_value_i = one_body_ref_diel_values.get(label1, 0.0) or 0.0
        one_body_value_j = one_body_ref_diel_values.get(label2, 0.0) or 0.0

        adjusted_value = (two_body_diel_value_to_use - one_body_value_i - one_body_value_j) * conversion_factor

        if label1 != label2:
            ref_diel_matrix[label1 - 1, label2 - 1] = adjusted_value
            ref_diel_matrix[label2 - 1, label1 - 1] = adjusted_value
        else:
            ref_diel_matrix[label1 - 1, label2 - 1] = np.nan

    for i in range(1, n):
        for j in range(i):
            ref_diel_matrix[i, j] = 0

    ref_diel_int_energy = np.nansum(ref_diel_matrix)

    if ref_diel_int_energy != 0:
        df = pd.DataFrame(ref_diel_matrix, index=range(1, n + 1), columns=range(1, n + 1))
        output_file_path = os.path.join(normalized_LEDAW_output_path_two_body, 'SOLV-fp.xlsx')
        normalized_output_file_path = normalize_path(output_file_path)
        os.makedirs(os.path.dirname(normalized_output_file_path), exist_ok=True)

        writer_args = {'engine': 'openpyxl'}
        if os.path.exists(normalized_output_file_path):
            writer_args['mode'] = 'a'
            writer_args['if_sheet_exists'] = 'replace'
        else:
            writer_args['mode'] = 'w'

        with pd.ExcelWriter(normalized_output_file_path, **writer_args) as writer:
            df.to_excel(writer, sheet_name='REF-DIEL')
        print(f"Two-body REF dielectric LED interaction energy matrix was written to {normalized_output_file_path}")

    return ref_diel_matrix, ref_diel_int_energy


def populate_twobody_ref_cds_matrices_non_bsse(one_body_orcaout_filenames, two_body_orcaout_directory, conversion_factor, two_body_labels, LEDAW_output_path_two_body):
    """Populate the two-body REF CDS interaction matrix and save it to an Excel file. Return the matrix and its sum."""

    normalized_LEDAW_output_path_two_body = normalize_path(LEDAW_output_path_two_body)

    cds_pattern = r"SMD CDS \(Gcds\)\s*:\s*([-+]?\d*\.\d+)"

    one_body_cds_values = extract_onebody_dielectric_or_cds_values(one_body_orcaout_filenames, cds_pattern)
    two_body_cds_values = extract_twobody_dielectric_or_cds_values(two_body_orcaout_directory, two_body_labels, cds_pattern)

    n = len(one_body_orcaout_filenames)
    ref_cds_matrix = np.zeros((n, n))

    all_cds_found = True
    for idx in range(1, n + 1):
        if one_body_cds_values.get(idx) is None:
            all_cds_found = False
    for filename in two_body_labels.keys():
        if two_body_cds_values.get(filename, (None, None, None))[2] is None:
            all_cds_found = False

    if not all_cds_found:
        print("Warning: At least one CDS component was not found in all files. Hence, the entire CDS contribution matrix will be set to zero.")
        ref_cds_matrix[:] = 0
        ref_cds_int_energy = np.nansum(ref_cds_matrix)
        if ref_cds_int_energy != 0:
            df = pd.DataFrame(ref_cds_matrix, index=range(1, n + 1), columns=range(1, n + 1))
            output_file_path = os.path.join(normalized_LEDAW_output_path_two_body, 'SOLV-fp.xlsx')
            normalized_output_file_path = normalize_path(output_file_path)
            os.makedirs(os.path.dirname(normalized_output_file_path), exist_ok=True)
            
            writer_args = {'engine': 'openpyxl'}
            if os.path.exists(normalized_output_file_path):
                writer_args['mode'] = 'a'
                writer_args['if_sheet_exists'] = 'replace'
            else:
                writer_args['mode'] = 'w'

            with pd.ExcelWriter(normalized_output_file_path, **writer_args) as writer:
                df.to_excel(writer, sheet_name='REF-CDS')
            print(f"Two-body REF CDS LED interaction energy matrix was written to {normalized_output_file_path}")
        return ref_cds_matrix, ref_cds_int_energy

    for filename, (label1, label2, two_body_cds_value) in two_body_cds_values.items():
        two_body_cds_value_to_use = two_body_cds_value or 0.0
        one_body_value_i = one_body_cds_values.get(label1, 0.0) or 0.0
        one_body_value_j = one_body_cds_values.get(label2, 0.0) or 0.0

        adjusted_value = (two_body_cds_value_to_use - one_body_value_i - one_body_value_j) * conversion_factor

        if label1 != label2:
            ref_cds_matrix[label1 - 1, label2 - 1] = adjusted_value
            ref_cds_matrix[label2 - 1, label1 - 1] = adjusted_value
        else:
            ref_cds_matrix[label1 - 1, label2 - 1] = np.nan

    for i in range(1, n):
        for j in range(i):
            ref_cds_matrix[i, j] = 0

    ref_cds_int_energy = np.nansum(ref_cds_matrix)

    if ref_cds_int_energy != 0:
        df = pd.DataFrame(ref_cds_matrix, index=range(1, n + 1), columns=range(1, n + 1))
        output_file_path = os.path.join(normalized_LEDAW_output_path_two_body, 'SOLV-fp.xlsx')
        normalized_output_file_path = normalize_path(output_file_path)
        os.makedirs(os.path.dirname(normalized_output_file_path), exist_ok=True)

        writer_args = {'engine': 'openpyxl'}
        if os.path.exists(normalized_output_file_path):
            writer_args['mode'] = 'a'
            writer_args['if_sheet_exists'] = 'replace'
        else:
            writer_args['mode'] = 'w'

        with pd.ExcelWriter(normalized_output_file_path, **writer_args) as writer:
            df.to_excel(writer, sheet_name='REF-CDS') 
        print(f"Two-body REF CDS LED interaction energy matrix was written to {normalized_output_file_path}")

    return ref_cds_matrix, ref_cds_int_energy


def populate_twobody_corr_dielectric_matrices_non_bsse(one_body_orcaout_filenames, two_body_orcaout_directory, conversion_factor, two_body_labels, LEDAW_output_path_two_body, method):
    """Populate the two-body correlation DIEL interaction matrix and save it to an Excel file. Return the sum of the correlation dielectric terms."""

    normalized_LEDAW_output_path_two_body = normalize_path(LEDAW_output_path_two_body)

    dielectric_pattern = r"C-PCM corr\. term \(included in E\(CORR\)\).*?([-+]?\d*\.\d+)"

    # Extract CORR dielectric values from one-body files
    one_body_corr_dielectric_values = extract_onebody_dielectric_or_cds_values(one_body_orcaout_filenames, dielectric_pattern)

    # Zero subsystem dielectric values if method is HFLD
    if method.lower() == "hfld":
        for k in one_body_corr_dielectric_values:
            one_body_corr_dielectric_values[k] = 0.0

    # Extract CORR dielectric values from two-body files
    two_body_corr_dielectric_values = extract_twobody_dielectric_or_cds_values(two_body_orcaout_directory, two_body_labels, dielectric_pattern)
    n = len(one_body_orcaout_filenames)
    corr_diel_matrix = np.zeros((n, n))

    for filename, (label1, label2) in two_body_labels.items():
        two_body_entry = two_body_corr_dielectric_values.get(filename)
        one_body_value_i = float(one_body_corr_dielectric_values.get(label1, 0.0) or 0.0)
        one_body_value_j = float(one_body_corr_dielectric_values.get(label2, 0.0) or 0.0)

        two_body_value = 0.0
        if two_body_entry is not None and len(two_body_entry) > 2 and two_body_entry[2] is not None:
            two_body_value = float(two_body_entry[2])

        adjusted_value = (two_body_value - one_body_value_i - one_body_value_j) * conversion_factor

        if label1 != label2:
            corr_diel_matrix[label1 - 1, label2 - 1] = adjusted_value
            corr_diel_matrix[label2 - 1, label1 - 1] = adjusted_value
        else:
            corr_diel_matrix[label1 - 1, label2 - 1] = np.nan

    for i in range(1, n):
        for j in range(i):
            corr_diel_matrix[i, j] = 0

    corr_diel_int_energy = np.nansum(corr_diel_matrix)

    if corr_diel_int_energy != 0:
        df = pd.DataFrame(corr_diel_matrix, index=range(1, n + 1), columns=range(1, n + 1))
        output_file_path = os.path.join(normalized_LEDAW_output_path_two_body, 'SOLV-fp.xlsx')
        normalized_output_file_path = normalize_path(output_file_path)
        if os.path.exists(normalized_output_file_path):
            with pd.ExcelWriter(normalized_output_file_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='CORR-DIEL')
        else:
            with pd.ExcelWriter(normalized_output_file_path, mode='w', engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='CORR-DIEL')

        print(f"Two-body CORR dielectric LED interaction energy matrix was written to {normalized_output_file_path}")

    return corr_diel_matrix, corr_diel_int_energy


def populate_twobody_total_solvation_matrices_non_bsse(one_body_orcaout_filenames, two_body_orcaout_directory, conversion_factor, two_body_labels, LEDAW_output_path_two_body, method):
    """Compute total solvation matrix for non-BSSE case by summing REF DIEL, REF CDS, and CORR DIEL contributions."""

    ref_cpcm_matrix, ref_cpcm_int_energy = populate_twobody_ref_diel_matrices_non_bsse(
        one_body_orcaout_filenames, two_body_orcaout_directory, conversion_factor, two_body_labels, LEDAW_output_path_two_body)

    ref_cds_matrix, ref_cds_int_energy = populate_twobody_ref_cds_matrices_non_bsse(
        one_body_orcaout_filenames, two_body_orcaout_directory, conversion_factor, two_body_labels, LEDAW_output_path_two_body)

    corr_diel_matrix, corr_diel_int_energy = populate_twobody_corr_dielectric_matrices_non_bsse(
        one_body_orcaout_filenames, two_body_orcaout_directory, conversion_factor, two_body_labels, LEDAW_output_path_two_body, method)

    total_solv_matrix = ref_cpcm_matrix + ref_cds_matrix + corr_diel_matrix
    solv_int_energy = ref_cpcm_int_energy + ref_cds_int_energy + corr_diel_int_energy

    output_file_path = os.path.join(normalize_path(LEDAW_output_path_two_body), 'SOLV-fp.xlsx')
    normalized_output_file_path = normalize_path(output_file_path)	
    df = pd.DataFrame(total_solv_matrix, index=range(1, total_solv_matrix.shape[0] + 1), columns=range(1, total_solv_matrix.shape[1] + 1))
    if os.path.exists(normalized_output_file_path):
        with pd.ExcelWriter(normalized_output_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name='SOLV')
    else:
        with pd.ExcelWriter(normalized_output_file_path, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name='SOLV')

    print(f"Two-body TOTAL solvation LED interaction energy matrix was written to {normalized_output_file_path}")

    return total_solv_matrix, solv_int_energy


def populate_twobody_ref_diel_matrices_bsse(bsse_file_pair_map, two_body_orcaout_directory, conversion_factor, two_body_labels, LEDAW_output_path_two_body):
    """Populate the BSSE-corrected REF DIEL interaction matrix and save it to an Excel file. Returns the matrix and its sum."""

    normalized_LEDAW_output_path_two_body = normalize_path(LEDAW_output_path_two_body)
    normalized_two_body_orcaout_directory = normalize_path(two_body_orcaout_directory)

    dielectric_pattern = r"CPCM Dielectric\s*:\s*(-?\d+\.\d+)\s*Eh"

    # Determine matrix size (n) from the maximum label in two_body_labels
    n = max(max(pair) for pair in two_body_labels.values())
    ref_cpcm_matrix = np.zeros((n, n))

    # --- Collect all unique relevant ORCA output file paths for global check ---
    all_relevant_filepaths = set()
    # Add two-body files
    for filename in two_body_labels.keys():
        all_relevant_filepaths.add(os.path.join(normalized_two_body_orcaout_directory, filename))
    # Add one-body files from bsse_file_pair_map
    for pair_files in bsse_file_pair_map.values():
        all_relevant_filepaths.add(normalize_path(pair_files[0]))
        all_relevant_filepaths.add(normalize_path(pair_files[1]))

    # --- Pre-scan these files for presence of DIEL data ---
    all_diel_found = True
    for file_path in all_relevant_filepaths:
        if not os.path.isfile(file_path):
            all_diel_found = False # Missing file means missing DIEL
            continue # Cannot read content if file doesn't exist

        try:
            with open(file_path, 'r') as f:
                content = f.read()
                if not re.search(dielectric_pattern, content):
                    all_diel_found = False # DIEL pattern not found in this file
        except Exception: # Catch potential file reading errors
            all_diel_found = False

    # Decision flag for DIEL
    is_diel_issue = not all_diel_found # True if any DIEL component is missing globally

    # If any DIEL is missing, zero the entire matrix and return.
    if is_diel_issue:
        print("Warning: At least one Dielectric component was not found in required files for BSSE-corrected REF CPCM calculation. Hence, the entire REF CPCM matrix will be set to zero.")
        ref_cpcm_matrix[:] = 0.0
        ref_cpcm_int_energy = 0.0

        # Excel writing logic, consistent with original function
        df = pd.DataFrame(ref_cpcm_matrix, index=range(1, n + 1), columns=range(1, n + 1))
        output_file_path_full = os.path.join(normalized_LEDAW_output_path_two_body, 'SOLV-fp.xlsx')
        normalized_output_file_path = normalize_path(output_file_path_full)
        
        os.makedirs(os.path.dirname(normalized_output_file_path), exist_ok=True)
        
        writer_args = {'engine': 'openpyxl'}
        if os.path.exists(normalized_output_file_path):
            writer_args['mode'] = 'a'
            writer_args['if_sheet_exists'] = 'replace'
        else:
            writer_args['mode'] = 'w'

        with pd.ExcelWriter(normalized_output_file_path, **writer_args) as writer:
            df.to_excel(writer, sheet_name='REF-DIEL')
        print(f"BSSE-corrected REF CPCM matrix (zeroed) was written to {normalized_output_file_path} on sheet 'REF-DIEL'.")
        return ref_cpcm_matrix, ref_cpcm_int_energy

    # --- Main calculation loop: Populating the matrix ---
    for filename, (label1, label2) in two_body_labels.items():
        two_body_file_path = os.path.join(normalized_two_body_orcaout_directory, filename)
        
        diel_two_body = 0.0

        # Read two-body file and extract dielectric value
        if os.path.isfile(two_body_file_path):
            try:
                with open(two_body_file_path, 'r') as f:
                    content = f.read()
                    match_diel = re.search(dielectric_pattern, content)
                    if match_diel:
                        diel_two_body = float(match_diel.group(1))
            except Exception:
                pass # If any error reading file, value remains 0.0

        # Look up and read one-body files for BSSE correction
        file_pair = bsse_file_pair_map.get((label1, label2)) or bsse_file_pair_map.get((label2, label1))
        if not file_pair:
            continue # If the one-body files for this pair aren't mapped, skip processing this pair

        onebody_file1, onebody_file2 = file_pair
        diel_one_body1 = diel_one_body2 = 0.0

        # Read one-body file 1
        if os.path.exists(onebody_file1):
            try:
                with open(onebody_file1, 'r') as f1:
                    content1 = f1.read()
                    match1_diel = re.search(dielectric_pattern, content1)
                    if match1_diel:
                        diel_one_body1 = float(match1_diel.group(1))
            except Exception:
                pass

        # Read one-body file 2
        if os.path.exists(onebody_file2):
            try:
                with open(onebody_file2, 'r') as f2:
                    content2 = f2.read()
                    match2_diel = re.search(dielectric_pattern, content2)
                    if match2_diel:
                        diel_one_body2 = float(match2_diel.group(1))
            except Exception:
                pass
        
        adjusted_val = (diel_two_body - diel_one_body1 - diel_one_body2) * conversion_factor
        i, j = label1 - 1, label2 - 1

        if i != j:
            ref_cpcm_matrix[i, j] = adjusted_val
            ref_cpcm_matrix[j, i] = adjusted_val
        else:
            ref_cpcm_matrix[i, i] = np.nan # Diagonal elements are NaN for interaction energy

    # Set values below the diagonal to zero as per convention
    for i in range(1, n):
        for j in range(i):
            ref_cpcm_matrix[i, j] = 0

    # Compute ref_cpcm_int_energy as the sum of all elements in the matrix (ignoring NaNs and zeros below diagonal)
    ref_cpcm_int_energy = np.nansum(ref_cpcm_matrix)

    # Write the matrix to an Excel file
    # This ensures it's written even if the value is zero, unless specifically excluded by original logic
    df = pd.DataFrame(ref_cpcm_matrix, index=range(1, n + 1), columns=range(1, n + 1))
    output_file_path_full = os.path.join(normalized_LEDAW_output_path_two_body, 'SOLV-fp.xlsx')
    normalized_output_file_path = normalize_path(output_file_path_full)
    
    os.makedirs(os.path.dirname(normalized_output_file_path), exist_ok=True)
    
    writer_args = {'engine': 'openpyxl'}
    if os.path.exists(normalized_output_file_path):
        writer_args['mode'] = 'a'
        writer_args['if_sheet_exists'] = 'replace'
    else:
        writer_args['mode'] = 'w'

    with pd.ExcelWriter(normalized_output_file_path, **writer_args) as writer:
        df.to_excel(writer, sheet_name='REF-DIEL')
    print(f"BSSE-corrected REF CPCM matrix was written to {normalized_output_file_path} on sheet 'REF-DIEL'.")

    return ref_cpcm_matrix, ref_cpcm_int_energy


def populate_twobody_ref_cds_matrices_bsse(bsse_file_pair_map, two_body_orcaout_directory, conversion_factor, two_body_labels, LEDAW_output_path_two_body):
    """Populate the BSSE-corrected REF CDS interaction matrix and save it to an Excel file. Returns the matrix and its sum."""

    normalized_LEDAW_output_path_two_body = normalize_path(LEDAW_output_path_two_body)
    normalized_two_body_orcaout_directory = normalize_path(two_body_orcaout_directory)

    cds_pattern = r"SMD CDS \(Gcds\)\s*:\s*([-+]?\d*\.\d+)"

    # Determine matrix size (n) from the maximum label in two_body_labels
    n = max(max(pair) for pair in two_body_labels.values())
    ref_cds_matrix = np.zeros((n, n))

    # --- Collect all unique relevant ORCA output file paths for global check ---
    all_relevant_filepaths = set()
    # Add two-body files
    for filename in two_body_labels.keys():
        all_relevant_filepaths.add(os.path.join(normalized_two_body_orcaout_directory, filename))
    # Add one-body files from bsse_file_pair_map
    for pair_files in bsse_file_pair_map.values():
        all_relevant_filepaths.add(normalize_path(pair_files[0]))
        all_relevant_filepaths.add(normalize_path(pair_files[1]))

    # --- Pre-scan these files for presence of CDS data ---
    any_cds_found = False # True if at least one CDS value is found in any file
    all_cds_found = True  # True only if ALL CDS values are found in all files

    for file_path in all_relevant_filepaths:
        if not os.path.isfile(file_path):
            all_cds_found = False # Missing file means missing CDS
            continue # Cannot read content if file doesn't exist

        try:
            with open(file_path, 'r') as f:
                content = f.read()
                if re.search(cds_pattern, content):
                    any_cds_found = True # CDS pattern found in this file
                else:
                    all_cds_found = False # CDS pattern not found in this file
        except Exception: # Catch potential file reading errors
            all_cds_found = False

    # Decision flags for how to handle CDS
    # True if some CDS are found but not all (partial presence)
    is_cds_uniform_issue = any_cds_found and not all_cds_found 
    force_zero_all_cds_contribution = False

    if is_cds_uniform_issue:
        force_zero_all_cds_contribution = True
        print("Warning: CDS components were not found in all required files for BSSE-corrected REF CDS calculation (partial presence detected).")
        print("Hence, the entire REF CDS contribution will be set to zero.")
    elif not any_cds_found: # If no CDS found anywhere at all
        force_zero_all_cds_contribution = True
        print("Note: No CDS components were found in any required files. The REF CDS matrix will be zero.")

    # --- Main calculation loop: Populating the matrix ---
    for filename, (label1, label2) in two_body_labels.items():
        two_body_file_path = os.path.join(normalized_two_body_orcaout_directory, filename)
        
        cds_two_body = 0.0 

        # Read two-body file and extract CDS value
        if os.path.isfile(two_body_file_path):
            try:
                with open(two_body_file_path, 'r') as f:
                    content = f.read()
                    # CDS is used only if not forced to zero AND any CDS values were found initially.
                    if not force_zero_all_cds_contribution and any_cds_found:
                        match_cds = re.search(cds_pattern, content)
                        if match_cds:
                            cds_two_body = float(match_cds.group(1))
            except Exception:
                pass # If any error reading file, value remains 0.0

        # Look up and read one-body files for BSSE correction
        file_pair = bsse_file_pair_map.get((label1, label2)) or bsse_file_pair_map.get((label2, label1))
        if not file_pair:
            continue # If the one-body files for this pair aren't mapped, skip processing this pair

        onebody_file1, onebody_file2 = file_pair
        cds_one_body1 = cds_one_body2 = 0.0 

        # Read one-body file 1
        if os.path.exists(onebody_file1):
            try:
                with open(onebody_file1, 'r') as f1:
                    content1 = f1.read()
                    # Handle CDS for one-body file 1 based on flags
                    if not force_zero_all_cds_contribution and any_cds_found:
                        match1_cds = re.search(cds_pattern, content1)
                        if match1_cds:
                            cds_one_body1 = float(match1_cds.group(1))
            except Exception:
                pass

        # Read one-body file 2
        if os.path.exists(onebody_file2):
            try:
                with open(onebody_file2, 'r') as f2:
                    content2 = f2.read()
                    # Handle CDS for one-body file 2 based on flags
                    if not force_zero_all_cds_contribution and any_cds_found:
                        match2_cds = re.search(cds_pattern, content2)
                        if match2_cds:
                            cds_one_body2 = float(match2_cds.group(1))
            except Exception:
                pass
        
        adjusted_val = (cds_two_body - cds_one_body1 - cds_one_body2) * conversion_factor
        i, j = label1 - 1, label2 - 1

        if i != j:
            ref_cds_matrix[i, j] = adjusted_val
            ref_cds_matrix[j, i] = adjusted_val
        else:
            ref_cds_matrix[i, i] = np.nan # Diagonal elements are NaN for interaction energy

    # Set values below the diagonal to zero as per convention
    for i in range(1, n):
        for j in range(i):
            ref_cds_matrix[i, j] = 0

    # Compute ref_cds_int_energy as the sum of all elements in the matrix (ignoring NaNs and zeros below diagonal)
    ref_cds_int_energy = np.nansum(ref_cds_matrix)

    # Write the matrix to an Excel file
    df = pd.DataFrame(ref_cds_matrix, index=range(1, n + 1), columns=range(1, n + 1))
    output_file_path_full = os.path.join(normalized_LEDAW_output_path_two_body, 'SOLV-fp.xlsx')
    normalized_output_file_path = normalize_path(output_file_path_full)

    os.makedirs(os.path.dirname(normalized_output_file_path), exist_ok=True)
    
    writer_args = {'engine': 'openpyxl'}
    if os.path.exists(normalized_output_file_path):
        writer_args['mode'] = 'a'
        writer_args['if_sheet_exists'] = 'replace'
    else:
        writer_args['mode'] = 'w'

    with pd.ExcelWriter(normalized_output_file_path, **writer_args) as writer:
        df.to_excel(writer, sheet_name='REF-CDS')
    print(f"BSSE-corrected REF CDS matrix was written to {normalized_output_file_path} on sheet 'REF-CDS'.")

    return ref_cds_matrix, ref_cds_int_energy


def populate_twobody_corr_dielectric_matrices_bsse(bsse_file_pair_map, two_body_orcaout_directory, conversion_factor, two_body_labels, LEDAW_output_path_two_body, method):
    """Populate the BSSE-corrected correlation DIEL interaction matrix and save it to an Excel file. Return the sum of the correlation DIEL terms."""

    normalized_LEDAW_output_path_two_body = normalize_path(LEDAW_output_path_two_body)
    normalized_two_body_orcaout_directory = normalize_path(two_body_orcaout_directory)
    dielectric_pattern = r"C-PCM corr\. term \(included in E\(CORR\)\).*?([-+]?\d*\.\d+)"
    n = max(max(pair) for pair in two_body_labels.values())
    corr_diel_matrix = np.zeros((n, n))

    for filename, (label1, label2) in two_body_labels.items():
        file_path = os.path.join(normalized_two_body_orcaout_directory, filename)
        if not os.path.isfile(file_path):
            two_body_value = 0.0
        else:
            with open(file_path, 'r') as f:
                content = f.read()
                match = re.search(dielectric_pattern, content)
                two_body_value = float(match.group(1)) if match else 0.0

        file_pair = bsse_file_pair_map.get((label1, label2)) or bsse_file_pair_map.get((label2, label1))
        if not file_pair:
            continue

        onebody_file1, onebody_file2 = file_pair
        corr_diel_one_body1 = corr_diel_one_body2 = 0.0

        if os.path.exists(onebody_file1):
            with open(onebody_file1, 'r') as f1:
                content1 = f1.read()
                match1 = re.search(dielectric_pattern, content1)
                if match1:
                    corr_diel_one_body1 = float(match1.group(1))

        if os.path.exists(onebody_file2):
            with open(onebody_file2, 'r') as f2:
                content2 = f2.read()
                match2 = re.search(dielectric_pattern, content2)
                if match2:
                    corr_diel_one_body2 = float(match2.group(1))

        # Zero subsystem dielectric values if method is HFLD
        if method.lower() == "hfld":
            corr_diel_one_body1 = 0.0
            corr_diel_one_body2 = 0.0

        adjusted_val = (two_body_value - corr_diel_one_body1 - corr_diel_one_body2) * conversion_factor
        i, j = label1 - 1, label2 - 1

        if i != j:
            corr_diel_matrix[i, j] = adjusted_val
            corr_diel_matrix[j, i] = adjusted_val
        else:
            corr_diel_matrix[i, i] = np.nan

    for i in range(1, n):
        for j in range(i):
            corr_diel_matrix[i, j] = 0

    corr_diel_int_energy = np.nansum(corr_diel_matrix)

    if corr_diel_int_energy != 0:
        df = pd.DataFrame(corr_diel_matrix, index=range(1, n + 1), columns=range(1, n + 1))
        output_file_path = os.path.join(normalized_LEDAW_output_path_two_body, 'SOLV-fp.xlsx')
        normalized_output_file_path = normalize_path(output_file_path)
        with pd.ExcelWriter(normalized_output_file_path, mode='a' if os.path.exists(normalized_output_file_path) else 'w', if_sheet_exists='replace', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='CORR-DIEL')
        print(f"Two-body CORR dielectric LED interaction energy matrix was written to {normalized_output_file_path}")

    return corr_diel_matrix, corr_diel_int_energy


def populate_twobody_total_solvation_matrices_bsse(bsse_file_pair_map, two_body_orcaout_directory, conversion_factor, two_body_labels, LEDAW_output_path_two_body, method):
    """Compute total solvation matrix for BSSE case by summing REF CPCM, REF CDS, and CORR CPCM contribution sand save it to an Excel file."""

    ref_cpcm_matrix, ref_cpcm_int_energy = populate_twobody_ref_diel_matrices_bsse(bsse_file_pair_map, two_body_orcaout_directory, conversion_factor, two_body_labels, LEDAW_output_path_two_body)

    ref_cds_matrix, ref_cds_int_energy = populate_twobody_ref_cds_matrices_bsse(bsse_file_pair_map, two_body_orcaout_directory, conversion_factor, two_body_labels, LEDAW_output_path_two_body)

    corr_diel_matrix, corr_diel_int_energy = populate_twobody_corr_dielectric_matrices_bsse(bsse_file_pair_map, two_body_orcaout_directory, conversion_factor, two_body_labels, LEDAW_output_path_two_body, method)

    total_solv_matrix = ref_cpcm_matrix + ref_cds_matrix + corr_diel_matrix
    solv_int_energy = ref_cpcm_int_energy + ref_cds_int_energy + corr_diel_int_energy

    output_file_path = os.path.join(normalize_path(LEDAW_output_path_two_body), 'SOLV-fp.xlsx')
    normalized_output_file_path = normalize_path(output_file_path)	
    df = pd.DataFrame(total_solv_matrix, index=range(1, total_solv_matrix.shape[0] + 1), columns=range(1, total_solv_matrix.shape[1] + 1))
    if os.path.exists(normalized_output_file_path):
        with pd.ExcelWriter(normalized_output_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name='SOLV')
    else:
        with pd.ExcelWriter(normalized_output_file_path, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name='SOLV')

    print(f"Two-body TOTAL solvation LED interaction energy matrix was written to {normalized_output_file_path}")

    return total_solv_matrix, solv_int_energy


def populate_twobody_inter_matrices(one_body_orcaout_filenames, two_body_orcaout_directory, conversion_factor, two_body_labels=None, LEDAW_output_path_two_body=None):
    results = {}
    patterns.PATTERNS['corr_inter_full'] = (
        r"Interaction correlation for Fragments\s+(\d+)\s+and\s+(\d+):\s+[-]+\s+"
        r"Inter strong pairs\s+([-]?\d*\.\d+)\s+\(.*?\)\s+"
        r"(Inter triples\s+([-]?\d*\.\d+)\s+\(.*?\)\s+)?"
        r"Inter weak pairs\s+([-]?\d*\.\d+)\s+\(.*?\)\s+[-]+\s+Total interaction"
    )

    property_mapping = {
        'Electrostat': {'pattern': 'ref_inter', 'group': 3},
        'Exchange': {'pattern': 'ref_inter', 'group': 4},
        'Inter SP': {'pattern': 'corr_inter_full', 'group': 3},
        'Inter T': {'pattern': 'corr_inter_full', 'group': 5},
        'Inter WP': {'pattern': 'corr_inter_full', 'group': 6},
        'Disp SP': {'pattern': 'dispersion_strong_pairs', 'group': 3},
    }

    # Normalize the directory paths
    normalized_two_body_orcaout_directory = normalize_path(two_body_orcaout_directory)
    normalized_LEDAW_output_path_two_body = normalize_path(LEDAW_output_path_two_body)

    # Initialize matrices and populate them based on extracted data
    for sheet_name, props in property_mapping.items():
        n = max(max(pair) for pair in two_body_labels.values())
        matrix = np.zeros((n, n))
        pattern = patterns.PATTERNS[props['pattern']]
        group_num = props['group']

        for filename, (label1, label2) in two_body_labels.items():
            file_path = os.path.join(normalized_two_body_orcaout_directory, filename)
            with open(file_path, 'r') as file:
                content = file.read()
                match = re.search(pattern, content)
                if match and len(match.groups()) >= group_num:
                    value = match.group(group_num)
                    if value:
                        value = value.strip()
                        numeric_value = float(value) * conversion_factor
                        i, j = label1 - 1, label2 - 1
                        if i < j:
                            matrix[i, j] = numeric_value
                        elif j < i:
                            matrix[j, i] = numeric_value

        df = pd.DataFrame(matrix, index=range(1, n+1), columns=range(1, n+1))
        results[sheet_name] = df

    # Save the results to an Excel file
    output_file_path = os.path.join(normalized_LEDAW_output_path_two_body, 'INTER.xlsx')
    normalized_output_file_path = normalize_path(output_file_path)
    with pd.ExcelWriter(normalized_output_file_path) as writer:
        for sheet_name, df in results.items():
            df.to_excel(writer, sheet_name=sheet_name)

    print(f"Two-body interfragment LED interaction energy matrices were written to {normalized_output_file_path}")


def extract_ref_minus_solvent_contr(content, patterns):
    """Fallback logic to extract REF or Total energy minus implicit solvation contribution (CPCM Dielectric plus SMD CDS)"""
    
    # Extract CPCM Dielectric value
    diel_match = re.search(r"CPCM Dielectric\s*:\s*([-+]?\d*\.\d+|\d+)", content)
    diel = float(diel_match.group(1)) if diel_match else 0.0

    # Extract SMD CDS value
    cds_match = re.search(r"SMD CDS \(Gcds\)\s*:\s*([-+]?\d*\.\d+)", content)
    cds = float(cds_match.group(1)) if cds_match else 0.0

    # Calculate total correction from both terms
    total_ref_solv_correction = diel + cds

    # Subtract combined correction from the reference energy (pattern_e0)
    ref_match = re.search(patterns.pattern_e0, content)
    if ref_match:
        raw_ref = float(ref_match.group(1))
        return raw_ref - total_ref_solv_correction

    # If reference energy not found, subtract combined correction from the Total Energy pattern
    total_match = re.search(r"Total Energy\s*:\s*([-+]?\d*\.\d+|\d+)", content)
    if total_match:
        return float(total_match.group(1)) - total_ref_solv_correction 

    return 0.0 # If no relevant energy match, assume zero


def extract_onebody_values(one_body_filenames, patterns, method):
    """Extract REF, SP, WP, T values from one-body files using patterns, with implicit solvation fallback logic."""
    one_body_values = {
        'ref': {}, 'strong_corr': {}, 'weak_corr': {}, 'triples_corr': {}, 'cpcm_corr': {}
    }

    for idx, filename in enumerate(one_body_filenames, start=1):
        label = idx
        if filename is None:
            continue  # Skip missing files

        try:
            with open(filename, 'r') as file:
                content = file.read()

                # --- REF with implicit solvation fallback ---
                ref_energy = extract_ref_minus_solvent_contr(content, patterns)
                one_body_values['ref'][label] = ref_energy

                # --- Strong pairs and DIEL correction ---
                strong_corr_match = re.search(patterns.pattern_strong_corr, content)
                if strong_corr_match:
                    strong_corr_value = float(strong_corr_match.group(1))

                    # Subtract C-PCM correlation term from SP if present and store it
                    cpcm_corr_match = re.search(r"C-PCM corr\. term \(included in E\(CORR\)\).*?([-+]?\d*\.\d+)", content)
                    if cpcm_corr_match:
                        cpcm_corr_value = float(cpcm_corr_match.group(1))
                        strong_corr_value -= cpcm_corr_value
                        one_body_values['cpcm_corr'][label] = cpcm_corr_value

                    one_body_values['strong_corr'][label] = strong_corr_value

                # --- Weak pairs correction ---
                weak_corr_match = re.search(patterns.pattern_weak_corr, content)
                if weak_corr_match:
                    one_body_values['weak_corr'][label] = float(weak_corr_match.group(1))

                # --- Triples correction ---
                triples_corr_match = re.search(patterns.pattern_triples_corr, content)
                if triples_corr_match:
                    one_body_values['triples_corr'][label] = float(triples_corr_match.group(1))

        except FileNotFoundError:
            raise FileNotFoundError(f"File {filename} does not exist. Please check the path.")
        except Exception as e:
            raise RuntimeError(f"An error occurred while processing file {filename}: {e}")

    return one_body_values


def extract_twobody_values(two_body_directory, two_body_labels, pattern_key, patterns):
    """Process two-body files and extract values based on the provided pattern key."""
    extracted_values = {}

    for filename in os.listdir(two_body_directory):
        file_path = os.path.join(two_body_directory, filename)
        
        # Ensure the file exists and is relevant to the two_body_labels before processing
        if not os.path.exists(file_path):
            print(f"Warning: File not found: {file_path}. Skipping.")
            continue
        if filename not in two_body_labels:
            continue

        # Get labels early as they are needed for the intra_ref_alt fallback
        label1, label2 = two_body_labels[filename]

        with open(file_path, 'r', encoding='utf-8') as file: # Added encoding for robustness
            content = file.read()

        value1 = 0.0
        value2 = 0.0
        value3 = 0.0 

        if pattern_key == 'intra_ref':
            # First, find matches using the primary 'intra_ref' pattern
            matches = re.findall(patterns.PATTERNS['intra_ref'], content)

            if matches:
                # If primary pattern found, extract values directly
                value1 = float(matches[0][0].strip())
                # Handle absence of second and third groups gracefully
                value2 = float(matches[0][1].strip()) if len(matches[0]) > 1 else 0.0
                value3 = float(matches[0][2].strip()) if len(matches[0]) > 2 else 0.0
            else:
                # If 'intra_ref' not found, fall back to 'intra_ref_alt'
                alt_matches = re.findall(patterns.PATTERNS['intra_ref_alt'], content)

                if alt_matches:
                    # Initialize temporary variables to store found values for label1 and label2
                    found_val1_for_label1 = 0.0
                    found_val2_for_label2 = 0.0
                    label1_energy_found = False
                    label2_energy_found = False

                    for fragment_num_str, total_energy_str in alt_matches:
                        try:
                            fragment_num = int(fragment_num_str)
                            energy = float(total_energy_str)

                            if fragment_num == label1:
                                found_val1_for_label1 = energy
                                label1_energy_found = True
                            if fragment_num == label2:
                                found_val2_for_label2 = energy
                                label2_energy_found = True
                        except ValueError:
                            continue # Skip to next match if parsing fails

                    # Assign the found energies to value1 and value2
                    if label1_energy_found:
                        value1 = found_val1_for_label1
                    if label2_energy_found:
                        value2 = found_val2_for_label2

        else:
            # Original logic for other pattern_keys (SP, WP, T, Singles)
            pattern = patterns.PATTERNS[pattern_key]
            matches = re.findall(pattern, content)

            if matches:
                value1 = float(matches[0][0].strip())
                value2 = float(matches[0][1].strip()) if len(matches[0]) > 1 else 0.0
                value3 = float(matches[0][2].strip()) if len(matches[0]) > 2 else 0.0
        
        extracted_values[filename] = (label1, label2, value1, value2, value3)

    return extracted_values


def populate_twobody_elprep_matrices_non_bsse(one_body_orcaout_filenames, two_body_orcaout_directory, conversion_factor, method, two_body_labels, LEDAW_output_path_two_body):
    """Populate non-BSSE ELPREP matrices and return individual (target, partner) diagonal contributions for SOLV-STD."""

    normalized_two_body_orcaout_directory = normalize_path(two_body_orcaout_directory)
    normalized_LEDAW_output_path_two_body = normalize_path(LEDAW_output_path_two_body)

    one_body_values = extract_onebody_values(one_body_orcaout_filenames, patterns, method)

    properties = {
        'REF': 'intra_ref',
        'SP': 'intra_strong_pairs',
        'WP': 'intra_weak_pairs',
        'T': 'intra_triples',
        'Singles': 'singles_contribution'
    }

    n = len(one_body_orcaout_filenames)
    results = {}
    nonaggregated_el_prep = {}

    for prop_name, pattern_key in properties.items():
        two_body_values = extract_twobody_values(
            normalized_two_body_orcaout_directory,
            two_body_labels,
            pattern_key,
            patterns
        )
        matrix = np.zeros((n, n))

        one_body_key = {
            'REF': 'ref',
            'SP': 'strong_corr',
            'WP': 'weak_corr',
            'T': 'triples_corr'
        }.get(prop_name, None)  # Singles not in one-body files, default to 0

        for _, (label1, label2, val1, val2, _) in two_body_values.items():
            if one_body_key:
                val1_1b = one_body_values.get(one_body_key, {}).get(label1, 0.0)
                val2_1b = one_body_values.get(one_body_key, {}).get(label2, 0.0)
            else:
                val1_1b = 0.0  # Singles
                val2_1b = 0.0

            adj1 = (val1 - val1_1b) * conversion_factor
            adj2 = (val2 - val2_1b) * conversion_factor
            total = adj1 + adj2

            i, j = label1 - 1, label2 - 1
            matrix[i, i] += adj1
            matrix[j, j] += adj2
            if i != j:
                matrix[min(i, j), max(i, j)] = total

            # Accumulate per-pair diagonal contributions
            method_lower = method.lower()
            if method_lower == 'hfld':
                if prop_name == 'REF':
                    nonaggregated_el_prep[(label1, label2)] = adj1
                    nonaggregated_el_prep[(label2, label1)] = adj2
            elif 'dlpno' in method_lower:
                # For DLPNO-CCSD or DLPNO-CCSD(T), sum all five components
                nonaggregated_el_prep[(label1, label2)] = nonaggregated_el_prep.get((label1, label2), 0.0) + adj1
                nonaggregated_el_prep[(label2, label1)] = nonaggregated_el_prep.get((label2, label1), 0.0) + adj2

        results[prop_name] = pd.DataFrame(matrix, index=range(1, n + 1), columns=range(1, n + 1))

    elprep_file = os.path.join(normalized_LEDAW_output_path_two_body, 'ELPREP.xlsx')
    normalized_elprep_file = normalize_path(elprep_file) 
    with pd.ExcelWriter(normalized_elprep_file) as writer:
        for sheet_name, df in results.items():
            df.to_excel(writer, sheet_name=sheet_name)

    print(f"Two-body LED electronic preparation matrices written to '{normalized_elprep_file}'")
    
    return nonaggregated_el_prep


def populate_twobody_elprep_matrices_bsse(one_body_orcaout_filenames, two_body_orcaout_directory, conversion_factor, method, two_body_labels, bsse_file_pair_map, LEDAW_output_path_two_body):
    """Populate BSSE-corrected ELPREP matrices and return individual (target, partner) diagonal contributions for SOLV-STD."""

    normalized_two_body_orcaout_directory = normalize_path(two_body_orcaout_directory)
    normalized_LEDAW_output_path_two_body = normalize_path(LEDAW_output_path_two_body)

    n = max(max(pair) for pair in two_body_labels.values())
    results = {}
    nonaggregated_el_prep = {}

    properties = ['REF', 'SP', 'WP', 'T', 'Singles']

    for prop_name in properties:
        matrix = np.zeros((n, n))

        for filename, (label1, label2) in two_body_labels.items():
            file_path = os.path.join(normalized_two_body_orcaout_directory, filename)
            with open(file_path, 'r') as f:
                content = f.read()

            val1 = val2 = 0.0 # Initialize val1 and val2 for the current file
            if prop_name == 'REF':
                matches = re.findall(patterns.PATTERNS['intra_ref'], content)

                if matches:
                    val1 = float(matches[0][0])
                    val2 = float(matches[0][1]) if len(matches[0]) > 1 and matches[0][1].strip() else 0.0
                else:
                    alt_matches = re.findall(patterns.PATTERNS['intra_ref_alt'], content)

                    # Initialize temporary variables to store found values for label1 and label2
                    found_val1_for_label1 = 0.0
                    found_val2_for_label2 = 0.0
                    
                    # Flags to track if values were actually found
                    label1_energy_found = False
                    label2_energy_found = False

                    if alt_matches:
                        for fragment_num_str, total_energy_str in alt_matches:
                            fragment_num = int(fragment_num_str)
                            energy = float(total_energy_str)

                            if fragment_num == label1:
                                found_val1_for_label1 = energy
                                label1_energy_found = True
                            if fragment_num == label2:
                                found_val2_for_label2 = energy
                                label2_energy_found = True
                        
                        # Assign to val1 and val2 only if the respective fragment energy was found
                        if label1_energy_found:
                            val1 = found_val1_for_label1
                        if label2_energy_found:
                            val2 = found_val2_for_label2
                    else:
                        pass
            elif prop_name == 'SP':
                matches = re.findall(r"Intra strong pairs\s+([-]?\d*\.\d+)(?:\s+([-]?\d*\.\d+))?\s+sum=", content)
                if matches:
                    val1 = float(matches[0][0])
                    val2 = float(matches[0][1]) if len(matches[0]) > 1 else 0.0
            elif prop_name == 'WP':
                matches = re.findall(r"Intra weak pairs\s+([-]?\d*\.\d+)(?:\s+([-]?\d*\.\d+))?\s+sum=", content)
                if matches:
                    val1 = float(matches[0][0])
                    val2 = float(matches[0][1]) if len(matches[0]) > 1 else 0.0
            elif prop_name == 'T':
                matches = re.findall(r"Intra triples\s+([-]?\d*\.\d+)(?:\s+([-]?\d*\.\d+))?\s+sum=", content)
                if matches:
                    val1 = float(matches[0][0])
                    val2 = float(matches[0][1]) if len(matches[0]) > 1 else 0.0
            elif prop_name == 'Singles':
                matches = re.findall(r"Singles contribution\s+([-]?\d*\.\d+)(?:\s+([-]?\d*\.\d+))?\s+sum=", content)
                if matches:
                    val1 = float(matches[0][0])
                    val2 = float(matches[0][1]) if len(matches[0]) > 1 else 0.0

            one_body_val1 = one_body_val2 = 0.0
            if (label1, label2) in bsse_file_pair_map:
                file_for_label1, file_for_label2 = bsse_file_pair_map[(label1, label2)]
            elif (label2, label1) in bsse_file_pair_map:
                file_for_label2, file_for_label1 = bsse_file_pair_map[(label2, label1)]
            else:
                continue

            for idx, obfile in enumerate([file_for_label1, file_for_label2]):
                if not os.path.exists(obfile):
                    continue

                with open(obfile, 'r') as f:
                    ob_content = f.read()

                diel = 0.0
                diel_match = re.search(r"CPCM Dielectric\s*:\s*([-+]?\d*\.\d+|\d+)", ob_content)
                if diel_match:
                    diel = float(diel_match.group(1))

                cds = 0.0
                cds_match = re.search(r"SMD CDS \(Gcds\)\s*:\s*([-+]?\d*\.\d+|\d+)", ob_content) 

                if cds_match:
                    cds = float(cds_match.group(1))

                if prop_name == 'REF':
                    e0_match = re.search(patterns.pattern_e0, ob_content)
                    total_match = re.search(r"Total Energy\s*:\s*([-+]?\d*\.\d+|\d+)", ob_content)
                    ref_val = 0.0
                    if e0_match:
                        ref_val = float(e0_match.group(1)) - diel - cds
                    elif total_match:
                        ref_val = float(total_match.group(1)) - diel - cds

                    if idx == 0:
                        one_body_val1 = ref_val
                    else:
                        one_body_val2 = ref_val

                elif prop_name == 'SP':
                    match = re.search(patterns.pattern_strong_corr, ob_content)
                    if match:
                        strong_corr_value = float(match.group(1))
                       
                        # Subtract CPCM correlation if present
                        cpcm_corr_match = re.search(r"C-PCM corr\. term \(included in E\(CORR\)\).*?([-+]?\d*\.\d+)", ob_content)
                        if cpcm_corr_match:
                            cpcm_corr_value = float(cpcm_corr_match.group(1))
                            strong_corr_value -= cpcm_corr_value
                
                        if idx == 0:
                            one_body_val1 = strong_corr_value
                        else:
                            one_body_val2 = strong_corr_value
                elif prop_name == 'WP':
                    match = re.search(patterns.pattern_weak_corr, ob_content)
                    if match:
                        if idx == 0:
                            one_body_val1 = float(match.group(1))
                        else:
                            one_body_val2 = float(match.group(1))
                elif prop_name == 'T':
                    match = re.search(patterns.pattern_triples_corr, ob_content)
                    if match:
                        if idx == 0:
                            one_body_val1 = float(match.group(1))
                        else:
                            one_body_val2 = float(match.group(1))
                elif prop_name == 'Singles':
                    one_body_val1 = one_body_val2 = 0.0

            adj1 = (val1 - one_body_val1) * conversion_factor
            adj2 = (val2 - one_body_val2) * conversion_factor
            total_val = adj1 + adj2

            i, j = label1 - 1, label2 - 1
            matrix[i, i] += adj1
            matrix[j, j] += adj2
            if i != j:
                matrix[min(i, j), max(i, j)] = total_val

            if prop_name == 'REF' and method.lower() == 'hfld':
                nonaggregated_el_prep[(label1, label2)] = adj1
                nonaggregated_el_prep[(label2, label1)] = adj2
            elif method.lower().startswith('dlpno'):
                nonaggregated_el_prep[(label1, label2)] = nonaggregated_el_prep.get((label1, label2), 0.0) + adj1
                nonaggregated_el_prep[(label2, label1)] = nonaggregated_el_prep.get((label2, label1), 0.0) + adj2

        results[prop_name] = pd.DataFrame(matrix, index=range(1, n + 1), columns=range(1, n + 1))

    elprep_file = os.path.join(normalized_LEDAW_output_path_two_body, 'ELPREP.xlsx')
    normalized_elprep_file = normalize_path(elprep_file) 
    with pd.ExcelWriter(normalized_elprep_file) as writer:
        for sheet_name, df in results.items():
            df.to_excel(writer, sheet_name=sheet_name)

    print(f"Two-body LED electronic preparation matrices written to '{normalized_elprep_file}'")

    return nonaggregated_el_prep


def extract_elprep_diagonals(elprep_file_path):
    """Extract diagonals from each sheet in the ELPREP.xlsx file."""
    elprep_sheets = pd.read_excel(elprep_file_path, sheet_name=None, index_col=0)
    diagonals = {}
    for sheet_name, df in elprep_sheets.items():
        diagonal = np.diag(df)
        diagonals[sheet_name] = diagonal
    return diagonals


def combine_diagonals(diagonals, method):
    """Combine diagonals according to the specified method."""
    
    # Calculate REF diagonals
    ref_diagonal = diagonals.get('REF', np.zeros_like(next(iter(diagonals.values()))))

    combined_diagonals = {}

    if method.lower() == 'dlpno-ccsd':
        c_ccsd_diagonal = (
            diagonals.get('SP', np.zeros_like(ref_diagonal)) +
            diagonals.get('WP', np.zeros_like(ref_diagonal)) +
            diagonals.get('Singles', np.zeros_like(ref_diagonal))
        )

        total_diagonal = c_ccsd_diagonal + ref_diagonal

        combined_diagonals['REF'] = ref_diagonal
        combined_diagonals['C-CCSD'] = c_ccsd_diagonal
        combined_diagonals['TOTAL'] = total_diagonal

    elif method.lower() == 'dlpno-ccsd(t)':
        c_ccsd_t_diagonal = (
            diagonals.get('SP', np.zeros_like(ref_diagonal)) +
            diagonals.get('WP', np.zeros_like(ref_diagonal)) +
            diagonals.get('T', np.zeros_like(ref_diagonal)) +
            diagonals.get('Singles', np.zeros_like(ref_diagonal))
        )

        total_ccsd_t_diagonal = c_ccsd_t_diagonal + ref_diagonal

        combined_diagonals['REF'] = ref_diagonal
        combined_diagonals['C-CCSD(T)'] = c_ccsd_t_diagonal
        combined_diagonals['TOTAL'] = total_ccsd_t_diagonal

    elif method.lower() == 'hfld':
        combined_diagonals['REF'] = ref_diagonal
        combined_diagonals['TOTAL'] = ref_diagonal

    else:
        raise ValueError("Unknown method provided.")

    return combined_diagonals


def set_diag_belowdiag_nan(df):
    """Set all diagonal and below-diagonal elements to NaN for Electrostat and Exchange matrices."""
    df = df.astype(float)
    df = df.copy() 
    for i in range(df.shape[0]):
        for j in range(i+1):
            df.iat[i, j] = np.nan
    return df


def set_belowdiag_nan(df):
    """Set the values below the diagonal in the matrix to NaN."""
    df = df.astype(float) 
    mask = np.tril(np.ones(df.shape), k=-1).astype(bool)
    df.values[mask] = np.nan
    return df


def set_diag_to_nan_if_zero(df):
    """Set diagonal elements to NaN if they are all zero."""
    df = df.astype(float) 
    diagonal = np.diag(df)
    
    # Check if all diagonal elements are zero
    if np.all(diagonal == 0):
        np.fill_diagonal(df.values, np.nan)
    
    return df


def finalize_els_exch_matrices_for_writing(summary_sheets):
    """Finalize matrices by setting diagonals and below-diagonal elements to NaN where necessary."""
    for key in summary_sheets.keys():
        if key in ['Electrostat', 'Exchange']:
            summary_sheets[key] = set_diag_belowdiag_nan(summary_sheets[key])
        else:
            summary_sheets[key] = set_belowdiag_nan(summary_sheets[key])
            summary_sheets[key] = set_diag_to_nan_if_zero(summary_sheets[key])
    return summary_sheets


def calculate_twobody_standard_LED_summary_matrices(LEDAW_output_path_two_body, method):
    normalized_LEDAW_output_path_two_body = normalize_path(LEDAW_output_path_two_body)

    elprep_file = os.path.join(normalized_LEDAW_output_path_two_body, 'ELPREP.xlsx')
    inter_file = os.path.join(normalized_LEDAW_output_path_two_body, 'INTER.xlsx')
    summary_file = os.path.join(normalized_LEDAW_output_path_two_body, 'Summary_Standard_LED_matrices.xlsx')

    inter_sheets = pd.read_excel(inter_file, sheet_name=None, index_col=0)
    elprep_diagonals = extract_elprep_diagonals(elprep_file)
    combined_diagonals = combine_diagonals(elprep_diagonals, method)

    summary_sheets = {}
    summary_sheets['Electrostat'] = inter_sheets['Electrostat']
    summary_sheets['Exchange'] = inter_sheets['Exchange']

    ref_matrix = inter_sheets['Electrostat'] + inter_sheets['Exchange']
    np.fill_diagonal(ref_matrix.values, np.diag(ref_matrix) + combined_diagonals['REF'])
    summary_sheets['REF'] = ref_matrix

    if method.lower() == 'dlpno-ccsd':
        disp_ccsd = inter_sheets['Disp SP'] + inter_sheets['Inter WP']
        summary_sheets['Disp CCSD'] = set_diag_belowdiag_nan(disp_ccsd)
        inter_nondisp_ccsd = inter_sheets['Inter SP'] - inter_sheets['Disp SP']
        summary_sheets['Inter-NonDisp-C-CCSD'] = set_diag_belowdiag_nan(inter_nondisp_ccsd)

        c_ccsd_matrix = inter_sheets['Inter SP'] + inter_sheets['Inter WP']
        np.fill_diagonal(c_ccsd_matrix.values, np.diag(c_ccsd_matrix) + combined_diagonals['C-CCSD'])
        summary_sheets['C-CCSD'] = c_ccsd_matrix

    elif method.lower() == 'dlpno-ccsd(t)':
        with np.errstate(divide='ignore', invalid='ignore'):
            disp_t = np.divide(inter_sheets['Disp SP'] * inter_sheets['Inter T'], inter_sheets['Inter SP'])
            disp_t[np.isnan(disp_t)] = 0
        disp_t = disp_t.where(np.triu(np.ones(disp_t.shape), k=0).astype(bool))
        inter_sheets['Disp T'] = disp_t

        disp_ccsd_t = inter_sheets['Disp SP'] + inter_sheets['Inter WP'] + disp_t
        summary_sheets['Disp CCSD(T)'] = set_diag_belowdiag_nan(disp_ccsd_t)

        inter_nondisp_ccsd_t = inter_sheets['Inter SP'] + inter_sheets['Inter WP'] + inter_sheets['Inter T'] - disp_ccsd_t
        summary_sheets['Inter-NonDisp-C-CCSD(T)'] = set_diag_belowdiag_nan(inter_nondisp_ccsd_t)

        c_ccsd_t_matrix = inter_sheets['Inter SP'] + inter_sheets['Inter WP'] + inter_sheets['Inter T']
        np.fill_diagonal(c_ccsd_t_matrix.values, np.diag(c_ccsd_t_matrix) + combined_diagonals['C-CCSD(T)'])
        summary_sheets['C-CCSD(T)'] = c_ccsd_t_matrix

    elif method.lower() == 'hfld':
        disp_hfld = inter_sheets['Disp SP'] + inter_sheets['Inter WP']
        summary_sheets['Disp HFLD'] = set_diag_belowdiag_nan(disp_hfld)

    for name, diagonal in combined_diagonals.items():
        diag_df = pd.DataFrame(np.diag(diagonal), index=range(1, len(diagonal)+1), columns=range(1, len(diagonal)+1))
        base_name = name if name == 'REF' else f'C-{name}'
        if base_name in summary_sheets:
            summary_sheets[base_name] = summary_sheets[base_name].add(diag_df, fill_value=0)

    # Explicitly add correlation EL-PREP diagonals
    correlation_terms = ['SP', 'WP', 'T', 'Singles']
    correlation_diag = sum([elprep_diagonals.get(term, 0) for term in correlation_terms])
    corr_elprep_df = pd.DataFrame(
        np.diag(correlation_diag), 
        index=range(1, len(correlation_diag)+1), 
        columns=range(1, len(correlation_diag)+1)
    )

    if method.lower() == 'dlpno-ccsd':
        summary_sheets['C-CCSD'] += corr_elprep_df
    elif method.lower() == 'dlpno-ccsd(t)':
        summary_sheets['C-CCSD(T)'] += corr_elprep_df

    summary_sheets = finalize_els_exch_matrices_for_writing(summary_sheets)

    sheet_order = [
        'SOLV', 'REF', 'Electrostat', 'Exchange',
        'C-CCSD' if method.lower() == 'dlpno-ccsd' else 'C-CCSD(T)',
        'Disp CCSD' if method.lower() == 'dlpno-ccsd' else 'Disp CCSD(T)' if method.lower() == 'dlpno-ccsd(t)' else 'Disp HFLD',
        'Inter-NonDisp-C-CCSD' if method.lower() == 'dlpno-ccsd' else 'Inter-NonDisp-C-CCSD(T)'
    ]

    with pd.ExcelWriter(summary_file, engine='openpyxl') as writer:
        for sheet_name in sheet_order:
            if sheet_name in summary_sheets:
                summary_sheets[sheet_name].to_excel(writer, sheet_name=sheet_name)

    workbook = openpyxl.load_workbook(summary_file)
    for sheet_name in sheet_order:
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.row > cell.column and cell.value == 0:
                        cell.value = None
    workbook.save(summary_file)

    normalized_summary_file = normalize_path(summary_file)
    print(f"Standard LED two-body summary interaction energy matrices were written to '{normalized_summary_file}'")


def compute_and_write_solv_std(LEDAW_output_path_two_body, method, nonaggregated_el_prep):
    """
    Computes and writes the SOLV-STD matrix and its individual components (REF-DIEL, REF-CDS, CORR-DIEL) to SOLV-STD.xlsx,
    based on pairwise computation from SOLV-fp components. Also updates the Summary_Standard_LED_matrices.xlsx.
    """
    normalized_path = normalize_path(LEDAW_output_path_two_body)

    # Note: Using 'SOLV-fp.xlsx' as the filename. If it was intended to be
    # f"SOLV-fp_{method}.xlsx" as in a previous thought, please adjust here.
    # The previous traceback showed 'SOLV-fp.xlsx'
    solv_fp_file = os.path.join(normalized_path, 'SOLV-fp.xlsx') 
    summary_file = os.path.join(normalized_path, 'Summary_Standard_LED_matrices.xlsx')
    solv_std_file = os.path.join(normalized_path, 'SOLV-STD.xlsx')

    # This function must work only if SOLV-fp.xlsx exists.
    # If the file doesn't exist, or if the crucial 'SOLV' sheet is missing,
    # it will proceed with a simplified summary update (without SOLV components).
    
    # Flag to track if full SOLV-STD computation should proceed
    perform_full_solv_std = True

    # Try to read the main 'SOLV' sheet from SOLV-fp.xlsx
    try:
        if not os.path.exists(solv_fp_file) or os.path.getsize(solv_fp_file) == 0:
            raise FileNotFoundError # Treat missing/empty file as an error for full computation
        
        solv_fp_df = pd.read_excel(solv_fp_file, sheet_name='SOLV', index_col=0).fillna(0.0)
        n = solv_fp_df.shape[0] # Get dimension from successfully loaded SOLV sheet
        
    except (FileNotFoundError, ValueError): # Catch if file not found or 'SOLV' sheet is missing
        print(f"Warning: '{solv_fp_file}' not found or sheet 'SOLV' is missing/empty. "
              "Skipping full SOLV-STD computation. Only updating Summary_Standard_LED_matrices.xlsx.")
        perform_full_solv_std = False

    if not perform_full_solv_std:
        # This block is executed if SOLV-fp.xlsx is missing/empty or 'SOLV' sheet is not found.
        summary_sheets = pd.read_excel(summary_file, sheet_name=None, index_col=0)

        # Compute TOTAL based on existing sheets, excluding SOLV
        if method.lower() == 'dlpno-ccsd':
            total = summary_sheets['REF'] + summary_sheets['C-CCSD']
        elif method.lower() == 'dlpno-ccsd(t)':
            total = summary_sheets['REF'] + summary_sheets['C-CCSD(T)']
        elif method.lower() == 'hfld':
            disp_hfld = summary_sheets['Disp HFLD'].copy()
            np.fill_diagonal(disp_hfld.values, 0.0) # treat as 0
            total = summary_sheets['REF'] + disp_hfld
        else:
            raise ValueError(f"Unknown method: {method}")

        total = total.where(np.triu(np.ones(total.shape), k=0).astype(bool))
        summary_sheets['TOTAL'] = total

        # Define sheet order for writing, explicitly excluding SOLV for this case
        sheet_order = [
            'TOTAL', 'REF', 'Electrostat', 'Exchange',
            'C-CCSD' if method.lower() == 'dlpno-ccsd' else 'C-CCSD(T)',
            'Disp CCSD' if method.lower() == 'dlpno-ccsd' else 'Disp CCSD(T)' if method.lower() == 'dlpno-ccsd(t)' else 'Disp HFLD',
            'Inter-NonDisp-C-CCSD' if method.lower() == 'dlpno-ccsd' else 'Inter-NonDisp-C-CCSD(T)'
        ]

        with pd.ExcelWriter(summary_file, engine='openpyxl') as writer:
            for sheet_name in sheet_order:
                if sheet_name in summary_sheets:
                    summary_sheets[sheet_name].to_excel(writer, sheet_name=sheet_name)

        # Mask zeros below diagonal for TOTAL (and potentially others if logic changes for them)
        workbook = openpyxl.load_workbook(summary_file)
        if 'TOTAL' in workbook.sheetnames:
            sheet = workbook['TOTAL']
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.row > cell.column and cell.value == 0:
                        cell.value = None
        workbook.save(summary_file)

        print(f"'TOTAL' sheet written without SOLV to '{normalize_path(summary_file)}'.")
        return # Exit the function here if full computation is skipped

    # --- Proceed with full SOLV-STD computation (only if perform_full_solv_std is True) ---
    
    # Initialize component dataframes with zeros, or load if sheets exist
    # These will be used in calculations; if a sheet is missing, it contributes zeros.
    ref_diel_fp_df = pd.DataFrame(np.zeros((n, n)), index=solv_fp_df.index, columns=solv_fp_df.columns)
    ref_cds_fp_df = pd.DataFrame(np.zeros((n, n)), index=solv_fp_df.index, columns=solv_fp_df.columns)
    corr_diel_fp_df = pd.DataFrame(np.zeros((n, n)), index=solv_fp_df.index, columns=solv_fp_df.columns)

    # Flags to indicate which component sheets were successfully loaded
    has_ref_diel = False
    has_ref_cds = False
    has_corr_diel = False

    try:
        ref_diel_fp_df = pd.read_excel(solv_fp_file, sheet_name='REF-DIEL', index_col=0).fillna(0.0)
        has_ref_diel = True
    except ValueError:
        print(f"Warning: Sheet 'REF-DIEL' not found in '{solv_fp_file}'.")
    
    try:
        ref_cds_fp_df = pd.read_excel(solv_fp_file, sheet_name='REF-CDS', index_col=0).fillna(0.0)
        has_ref_cds = True
    except ValueError:
        print(f"Warning: Sheet 'REF-CDS' not found in '{solv_fp_file}'.")

    try:
        corr_diel_fp_df = pd.read_excel(solv_fp_file, sheet_name='CORR-DIEL', index_col=0).fillna(0.0)
        has_corr_diel = True
    except ValueError:
        print(f"Warning: Sheet 'CORR-DIEL' not found in '{solv_fp_file}'.")

    summary_sheets = pd.read_excel(summary_file, sheet_name=None, index_col=0)

    # Determine INT matrix (existing logic)
    if method.lower() == 'dlpno-ccsd':
        int_matrix = summary_sheets['REF'] + summary_sheets['C-CCSD']
    elif method.lower() == 'dlpno-ccsd(t)':
        int_matrix = summary_sheets['REF'] + summary_sheets['C-CCSD(T)']
    elif method.lower() == 'hfld':
        int_matrix = summary_sheets['REF'] + summary_sheets['Disp HFLD']
    else:
        raise ValueError(f"Unknown method: {method}")

    int_matrix = int_matrix.fillna(0.0) # Ensure INT matrix is filled for calculations

    solv_std_matrix = np.zeros((n, n))
    ref_diel_std_matrix = np.zeros((n, n))
    ref_cds_std_matrix = np.zeros((n, n))
    corr_diel_std_matrix = np.zeros((n, n))

    for i in range(n):
        for j in range(n):
            # Values for total SOLV-STD computation
            solv_fp_ij = solv_fp_df.iloc[i, j]
            solv_fp_ji = solv_fp_df.iloc[j, i]
            solv_fp_val = solv_fp_ij if solv_fp_ij != 0 else solv_fp_ji # Choose non-zero or upper triangle

            int_ij = int_matrix.iloc[i, j]
            int_ji = int_matrix.iloc[j, i]
            int_val = int_ij if int_ij != 0 else int_ji # Choose non-zero or upper triangle

            prep_ij = nonaggregated_el_prep.get((i + 1, j + 1), 0.0)
            prep_ji = nonaggregated_el_prep.get((j + 1, i + 1), 0.0)

            # Common denominator for all STD components
            denom = int_val + prep_ij + prep_ji

            if i == j: # Diagonal elements calculation
                solv_std_ii = 0.0
                ref_diel_std_ii = 0.0
                ref_cds_std_ii = 0.0
                corr_diel_std_ii = 0.0

                for k in range(n):
                    if k != i:
                        # Common values for k-th interaction with i
                        prep_ik = nonaggregated_el_prep.get((i + 1, k + 1), 0.0)
                        prep_ki = nonaggregated_el_prep.get((k + 1, i + 1), 0.0)
                        int_ik = int_matrix.iloc[i, k]
                        int_ki = int_matrix.iloc[k, i]
                        int_val_diag = int_ik if int_ik != 0 else int_ki
                        denom_diag = int_val_diag + prep_ik + prep_ki # Denominator for diagonal elements

                        # Total SOLV-STD diagonal part
                        solv_fp_ik = solv_fp_df.iloc[i, k]
                        solv_fp_ki = solv_fp_df.iloc[k, i]
                        solv_fp_val_diag = solv_fp_ik if solv_fp_ik != 0 else solv_fp_ki
                        num_solv_std_diag = solv_fp_val_diag * prep_ik
                        solv_std_ii += num_solv_std_diag / denom_diag if denom_diag != 0 else 0.0

                        # REF-DIEL-STD diagonal part (only if source sheet was present)
                        if has_ref_diel:
                            ref_diel_fp_ik = ref_diel_fp_df.iloc[i, k]
                            ref_diel_fp_ki = ref_diel_fp_df.iloc[k, i]
                            ref_diel_fp_val_diag = ref_diel_fp_ik if ref_diel_fp_ik != 0 else ref_diel_fp_ki
                            num_ref_diel_diag = ref_diel_fp_val_diag * prep_ik
                            ref_diel_std_ii += num_ref_diel_diag / denom_diag if denom_diag != 0 else 0.0

                        # REF-CDS-STD diagonal part (only if source sheet was present)
                        if has_ref_cds:
                            ref_cds_fp_ik = ref_cds_fp_df.iloc[i, k]
                            ref_cds_fp_ki = ref_cds_fp_df.iloc[k, i]
                            ref_cds_fp_val_diag = ref_cds_fp_ik if ref_cds_fp_ik != 0 else ref_cds_fp_ki
                            num_ref_cds_diag = ref_cds_fp_val_diag * prep_ik
                            ref_cds_std_ii += num_ref_cds_diag / denom_diag if denom_diag != 0 else 0.0

                        # CORR-DIEL-STD diagonal part (only if source sheet was present)
                        if has_corr_diel:
                            corr_diel_fp_ik = corr_diel_fp_df.iloc[i, k]
                            corr_diel_fp_ki = corr_diel_fp_df.iloc[k, i]
                            corr_diel_fp_val_diag = corr_diel_fp_ik if corr_diel_fp_ik != 0 else corr_diel_fp_ki
                            num_corr_diel_diag = corr_diel_fp_val_diag * prep_ik
                            corr_diel_std_ii += num_corr_diel_diag / denom_diag if denom_diag != 0 else 0.0

                solv_std_matrix[i, i] = solv_std_ii
                ref_diel_std_matrix[i, i] = ref_diel_std_ii
                ref_cds_std_matrix[i, i] = ref_cds_std_ii
                corr_diel_std_matrix[i, i] = corr_diel_std_ii

            elif i < j: # Off-diagonal elements calculation (upper triangle)
                # Total SOLV-STD
                num_solv_std = solv_fp_val * int_val
                value_solv_std = num_solv_std / denom if denom != 0 else 0.0
                solv_std_matrix[i, j] = value_solv_std
                solv_std_matrix[j, i] = value_solv_std # Symmetrical

                # REF-DIEL-STD (only if source sheet was present)
                if has_ref_diel:
                    ref_diel_fp_ij = ref_diel_fp_df.iloc[i, j]
                    ref_diel_fp_ji = ref_diel_fp_df.iloc[j, i]
                    ref_diel_fp_val = ref_diel_fp_ij if ref_diel_fp_ij != 0 else ref_diel_fp_ji
                    num_ref_diel_std = ref_diel_fp_val * int_val
                    value_ref_diel_std = num_ref_diel_std / denom if denom != 0 else 0.0
                    ref_diel_std_matrix[i, j] = value_ref_diel_std
                    ref_diel_std_matrix[j, i] = value_ref_diel_std # Symmetrical

                # REF-CDS-STD (only if source sheet was present)
                if has_ref_cds:
                    ref_cds_fp_ij = ref_cds_fp_df.iloc[i, j]
                    ref_cds_fp_ji = ref_cds_fp_df.iloc[j, i]
                    ref_cds_fp_val = ref_cds_fp_ij if ref_cds_fp_ij != 0 else ref_cds_fp_ji
                    num_ref_cds_std = ref_cds_fp_val * int_val
                    value_ref_cds_std = num_ref_cds_std / denom if denom != 0 else 0.0
                    ref_cds_std_matrix[i, j] = value_ref_cds_std
                    ref_cds_std_matrix[j, i] = value_ref_cds_std # Symmetrical

                # CORR-DIEL-STD (only if source sheet was present)
                if has_corr_diel:
                    corr_diel_fp_ij = corr_diel_fp_df.iloc[i, j]
                    corr_diel_fp_ji = corr_diel_fp_df.iloc[j, i]
                    corr_diel_fp_val = corr_diel_fp_ij if corr_diel_fp_ij != 0 else corr_diel_fp_ji
                    num_corr_diel_std = corr_diel_fp_val * int_val
                    value_corr_diel_std = num_corr_diel_std / denom if denom != 0 else 0.0
                    corr_diel_std_matrix[i, j] = value_corr_diel_std
                    corr_diel_std_matrix[j, i] = value_corr_diel_std # Symmetrical

    # Create DataFrames for writing to Excel
    solv_std_df = pd.DataFrame(solv_std_matrix, index=solv_fp_df.index, columns=solv_fp_df.columns)
    ref_diel_std_df = pd.DataFrame(ref_diel_std_matrix, index=solv_fp_df.index, columns=solv_fp_df.columns)
    ref_cds_std_df = pd.DataFrame(ref_cds_std_matrix, index=solv_fp_df.index, columns=solv_fp_df.columns)
    corr_diel_std_df = pd.DataFrame(corr_diel_std_matrix, index=solv_fp_df.index, columns=solv_fp_df.columns)

    # --- Write all components to SOLV-STD.xlsx ---
    os.makedirs(os.path.dirname(solv_std_file), exist_ok=True) # Ensure directory exists

    with pd.ExcelWriter(solv_std_file, engine='openpyxl', mode='w') as writer:
        # Only write sheets that were successfully processed (their source sheet existed)
        if has_ref_diel:
            ref_diel_std_df.to_excel(writer, sheet_name='REF-DIEL')

        if has_ref_cds:
            ref_cds_std_df.to_excel(writer, sheet_name='REF-CDS')

        if has_corr_diel:
            corr_diel_std_df.to_excel(writer, sheet_name='CORR-DIEL')

        # The 'SOLV' sheet is always written if perform_full_solv_std was True
        solv_std_df.to_excel(writer, sheet_name='SOLV')

    # Create SOLV sheet (upper triangle only) and update summary_file
    solv_summary_df = solv_std_df.where(np.triu(np.ones(solv_std_df.shape), k=0).astype(bool))
    summary_sheets['SOLV'] = solv_summary_df # Add SOLV to summary sheets

    # Compute TOTAL with SOLV contribution
    if method.lower() == 'dlpno-ccsd':
        total = summary_sheets['REF'] + summary_sheets['C-CCSD']
    elif method.lower() == 'dlpno-ccsd(t)':
        total = summary_sheets['REF'] + summary_sheets['C-CCSD(T)']
    elif method.lower() == 'hfld':
        disp_hfld = summary_sheets['Disp HFLD'].copy()
        np.fill_diagonal(disp_hfld.values, 0.0)
        total = summary_sheets['REF'] + disp_hfld
    else:
        raise ValueError(f"Unknown method: {method}")

    total += solv_summary_df # Add SOLV to the total
    total = total.where(np.triu(np.ones(total.shape), k=0).astype(bool))
    summary_sheets['TOTAL'] = total

    # Write updated summary file
    sheet_order_summary = [
        'TOTAL', 'SOLV', 'REF', 'Electrostat', 'Exchange',
        'C-CCSD' if method.lower() == 'dlpno-ccsd' else 'C-CCSD(T)',
        'Disp CCSD' if method.lower() == 'dlpno-ccsd' else 'Disp CCSD(T)' if method.lower() == 'dlpno-ccsd(t)' else 'Disp HFLD',
        'Inter-NonDisp-C-CCSD' if method.lower() == 'dlpno-ccsd' else 'Inter-NonDisp-C-CCSD(T)'
    ]

    with pd.ExcelWriter(summary_file, engine='openpyxl') as writer:
        for sheet_name in sheet_order_summary:
            if sheet_name in summary_sheets: # Ensure the sheet exists in the collected summary_sheets
                summary_sheets[sheet_name].to_excel(writer, sheet_name=sheet_name)

    # Mask zeros below diagonal as None for SOLV and TOTAL sheets in summary file
    workbook = openpyxl.load_workbook(summary_file)
    for sheet_name in ['SOLV', 'TOTAL']:
        if sheet_name in workbook.sheetnames: # Only process if sheet actually exists
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.row > cell.column and cell.value == 0:
                        cell.value = None
    workbook.save(summary_file)

    print(f"Updated '{normalize_path(summary_file)}' with SOLV and TOTAL sheets.")


def extract_upper_diagonal(ref_sheet):
    """Extract the upper-diagonal elements"""
    upper_diag = np.triu(ref_sheet, k=1)
    return pd.DataFrame(upper_diag, index=ref_sheet.index, columns=ref_sheet.columns)


def sum_corr_elprep_upperdiagonals(elprep_file_path, include_t=False):
    """Sum the upper-diagonal elements of correlation components of ELPREP."""
    
    # Load the specified sheets from ELPREP.xlsx
    sheet_names = ['SP', 'WP', 'Singles']
    if include_t:
        sheet_names.append('T')
        
    elprep_sheets = pd.read_excel(elprep_file_path, sheet_name=sheet_names, index_col=0)
    
    # Initialize the sum with SP sheet's upper diagonal elements
    summed_matrix = elprep_sheets['SP'].where(np.triu(np.ones(elprep_sheets['SP'].shape), k=1).astype(bool), 0)

    # Add WP sheet's upper diagonal elements
    if 'WP' in elprep_sheets:
        wp_upper_diag = elprep_sheets['WP'].where(np.triu(np.ones(elprep_sheets['WP'].shape), k=1).astype(bool), 0)
        summed_matrix += wp_upper_diag

    # Add Singles sheet's upper diagonal elements
    if 'Singles' in elprep_sheets:
        singles_upper_diag = elprep_sheets['Singles'].where(np.triu(np.ones(elprep_sheets['Singles'].shape), k=1).astype(bool), 0)
        summed_matrix += singles_upper_diag

    # Add T sheet's upper diagonal elements if required
    if include_t and 'T' in elprep_sheets:
        t_upper_diag = elprep_sheets['T'].where(np.triu(np.ones(elprep_sheets['T'].shape), k=1).astype(bool), 0)
        summed_matrix += t_upper_diag

    return summed_matrix


def calculate_twobody_fpLED_matrices(LEDAW_output_path_two_body, method):
    # Normalize the LEDAW_output_path_two_body
    normalized_LEDAW_output_path_two_body = normalize_path(LEDAW_output_path_two_body)

    # Ensure the output directory exists
    if not os.path.exists(normalized_LEDAW_output_path_two_body):
        os.makedirs(normalized_LEDAW_output_path_two_body)

    # File paths
    elprep_file = os.path.join(normalized_LEDAW_output_path_two_body, 'ELPREP.xlsx')
    summary_standard_file = os.path.join(normalized_LEDAW_output_path_two_body, 'Summary_Standard_LED_matrices.xlsx')
    solv_file = os.path.join(normalized_LEDAW_output_path_two_body, 'SOLV-fp.xlsx')
    summary_file = os.path.join(normalized_LEDAW_output_path_two_body, 'Summary_fp-LED_matrices.xlsx')
    
    # Load standard sheets from Summary_Standard_LED_matrices.xlsx
    standard_sheets = pd.read_excel(summary_standard_file, sheet_name=None, index_col=0)
    
    # Initialize the summary sheets dictionary
    summary_sheets = {}

    # Transfer Electrostat and Exchange sheets from the standard sheets
    summary_sheets['Electrostat'] = standard_sheets['Electrostat']
    summary_sheets['Exchange'] = standard_sheets['Exchange']

    # Load the REF sheet from ELPREP.xlsx
    elprep_sheets = pd.read_excel(elprep_file, sheet_name=['REF'], index_col=0)
    ref_sheet = elprep_sheets['REF']
    
    # Calculate REF-EL-PREP as the upper diagonal of the REF sheet
    summary_sheets['REF-EL-PREP'] = extract_upper_diagonal(ref_sheet)
    
    # Calculate REF by summing Electrostat, Exchange, and REF-EL-PREP
    summary_sheets['REF'] = summary_sheets['Electrostat'] + summary_sheets['Exchange'] + summary_sheets['REF-EL-PREP']
    
    # Calculate C-CCSD-EL-PREP or C-CCSD(T)-EL-PREP depending on the method
    include_t = 't' in method.lower()
    c_ccsd_elprep = sum_corr_elprep_upperdiagonals(elprep_file, include_t=include_t)
    elprep_key = 'C-CCSD-EL-PREP' if not include_t else 'C-CCSD(T)-EL-PREP'
    
    # Assign the summed matrix to the appropriate key
    summary_sheets[elprep_key] = c_ccsd_elprep
    
    # Transfer Disp and Inter-NonDisp sheets from the standard sheets
    if method.lower() == 'dlpno-ccsd':
        summary_sheets['Disp CCSD'] = standard_sheets['Disp CCSD']
        summary_sheets['Inter-NonDisp-C-CCSD'] = standard_sheets['Inter-NonDisp-C-CCSD']
        # Calculate C-CCSD
        summary_sheets['C-CCSD'] = summary_sheets['Disp CCSD'] + summary_sheets['Inter-NonDisp-C-CCSD'] + summary_sheets['C-CCSD-EL-PREP']
    elif method.lower() == 'dlpno-ccsd(t)':
        summary_sheets['Disp CCSD(T)'] = standard_sheets['Disp CCSD(T)']
        summary_sheets['Inter-NonDisp-C-CCSD(T)'] = standard_sheets['Inter-NonDisp-C-CCSD(T)']
        # Calculate C-CCSD(T)
        summary_sheets['C-CCSD(T)'] = summary_sheets['Disp CCSD(T)'] + summary_sheets['Inter-NonDisp-C-CCSD(T)'] + summary_sheets['C-CCSD(T)-EL-PREP']
    elif method.lower() == 'hfld':
        summary_sheets['Disp HFLD'] = standard_sheets['Disp HFLD']

    # Calculate CCSD-EL-PREP or CCSD(T)-EL-PREP as the sum of REF-EL-PREP and C-CCSD-EL-PREP (or C-CCSD(T)-EL-PREP)
    if method.lower() == 'dlpno-ccsd':
        summary_sheets['CCSD-EL-PREP'] = summary_sheets['REF-EL-PREP'] + summary_sheets['C-CCSD-EL-PREP']
    elif method.lower() == 'dlpno-ccsd(t)':
        summary_sheets['CCSD(T)-EL-PREP'] = summary_sheets['REF-EL-PREP'] + summary_sheets['C-CCSD(T)-EL-PREP']

    # Transfer SOLV sheet if it is not zero
    if os.path.exists(solv_file):
        solv_sheet = pd.read_excel(solv_file, sheet_name=None, index_col=0)
        solv_df = solv_sheet['SOLV']

        # Check if the SOLV sheet is not entirely zero
        if not np.all(solv_df.values == 0):
            # Set diagonal and lower-diagonal elements to NaN
            mask = np.tril(np.ones(solv_df.shape), k=0).astype(bool)
            solv_df = solv_df.mask(mask)
            summary_sheets['SOLV'] = solv_df

    # Calculate TOTAL
    if method.lower() == 'hfld':
        summary_sheets['TOTAL'] = summary_sheets['REF'] + summary_sheets['Disp HFLD']
        if 'SOLV' in summary_sheets:
            summary_sheets['TOTAL'] += summary_sheets['SOLV']
    elif method.lower() == 'dlpno-ccsd':
        summary_sheets['TOTAL'] = summary_sheets['REF'] + summary_sheets['C-CCSD']
        if 'SOLV' in summary_sheets:
            summary_sheets['TOTAL'] += summary_sheets['SOLV']
    elif method.lower() == 'dlpno-ccsd(t)':
        summary_sheets['TOTAL'] = summary_sheets['REF'] + summary_sheets['C-CCSD(T)']
        if 'SOLV' in summary_sheets:
            summary_sheets['TOTAL'] += summary_sheets['SOLV']

    # Set diagonal and below-diagonal elements to NaN for all matrices
    for sheet_name in summary_sheets:
        df = summary_sheets[sheet_name]
        # Ensure that the DataFrame is of float type so it can accept NaN values
        df = df.astype(float)
        # Create a mask for diagonal and below-diagonal elements
        mask = np.tril(np.ones(df.shape), k=0).astype(bool)
        # Set these elements to NaN
        df.values[mask] = np.nan
        # Update the summary_sheets with the modified DataFrame
        summary_sheets[sheet_name] = df

    # Ensure that the file is created and write the results to the Excel file
    with pd.ExcelWriter(summary_file, engine='openpyxl') as writer:
        sheet_order = [
            'TOTAL', 'SOLV', 'REF', 'Electrostat', 'Exchange', 'REF-EL-PREP',
            'C-CCSD' if method.lower() == 'dlpno-ccsd' else 'C-CCSD(T)',
            'Disp CCSD' if method.lower() == 'dlpno-ccsd' else 'Disp CCSD(T)' if method.lower() == 'dlpno-ccsd(t)' else 'Disp HFLD',
            'Inter-NonDisp-C-CCSD' if method.lower() == 'dlpno-ccsd' else 'Inter-NonDisp-C-CCSD(T)',
            'C-CCSD-EL-PREP' if method.lower() == 'dlpno-ccsd' else 'C-CCSD(T)-EL-PREP',
            'CCSD-EL-PREP' if method.lower() == 'dlpno-ccsd' else 'CCSD(T)-EL-PREP'
        ]
        for sheet_name in sheet_order:
            if sheet_name in summary_sheets:
                summary_sheets[sheet_name].to_excel(writer, sheet_name=sheet_name)

    normalized_summary_file = normalize_path(summary_file)
    print(f"fp-LED two-body summary interaction energy matrices were written to '{normalized_summary_file}'")


def cleanup_and_reindex_all_excels(LEDAW_output_path_two_body, reduced_relabel_mapping=None, LEDAW_output_path_nbody=None):
    """Remove redundant rows/columns (with only 0 or NaN in the upper triangle, typical in BSSE cases)
    from all relevant Excel files in the given two-body LEDAW output directory.
    Optionally, relabel + reorder all Excel sheets using a pre-normalized reduced_relabel_mapping,
    and move all interaction values to the upper triangle. Applies relabeling from N-body if available.
    Also applies specific NaN masking for SOLV-fp.xlsx and INTER.xlsx (diagonal and below)
    and other files (below diagonal).
    """

    files_to_process = {
        "Summary_Standard_LED_matrices.xlsx": None,
        "Summary_fp-LED_matrices.xlsx": None,
        "SOLV-fp.xlsx": None,
        "SOLV-STD.xlsx": None,
        "INTER.xlsx": None,
        "ELPREP.xlsx": None
    }

    reference_file = os.path.join(LEDAW_output_path_two_body, "Summary_Standard_LED_matrices.xlsx")
    if not os.path.exists(reference_file):
        print("Reference file for cleaning matrices not found.")
        return

    # Step 1: Identify rows/columns to drop (consider diagonal + upper triangle)
    reference_sheets = pd.read_excel(reference_file, sheet_name=None, index_col=0)
    matrix_shape = next(iter(reference_sheets.values())).shape
    n = matrix_shape[0]

    zero_indices = set(range(n))  # 0-based indices

    for df in reference_sheets.values():
        df = df.fillna(0.0)
        arr = df.to_numpy()
        for i in range(n):
            has_nonzero = arr[i, i] != 0 or any(arr[i, j] != 0 for j in range(i + 1, n))
            if has_nonzero:
                zero_indices.discard(i)

    rows_to_drop = sorted(zero_indices)
    cols_to_drop = sorted(zero_indices)

    # Step 2: Apply cleanup to all Excel files
    for filename, selected_sheets in files_to_process.items():
        full_path = os.path.join(LEDAW_output_path_two_body, filename)
        if not os.path.exists(full_path):
            continue

        sheets = pd.read_excel(full_path, sheet_name=selected_sheets, index_col=0)
        if isinstance(sheets, pd.DataFrame):
            sheets = {selected_sheets[0]: sheets}

        cleaned_sheets = {}
        for sheet_name, df in sheets.items():
            # Drop identified rows and columns
            df = df.drop(index=df.index[rows_to_drop], errors='ignore')
            df = df.drop(columns=df.columns[cols_to_drop], errors='ignore')

            # Reindex rows and columns consecutively from 1 to n (after cleanup)
            new_index = list(range(1, df.shape[0] + 1))
            df.index = new_index
            df.columns = new_index

            # Apply reduced relabel mapping if provided
            if reduced_relabel_mapping and len(df) == len(reduced_relabel_mapping):
                label_map = dict(zip(df.index, reduced_relabel_mapping))
                df = df.rename(index=label_map, columns=label_map)
                df = df.sort_index()
                df = df[sorted(df.columns)]  # Ensure columns sorted

                # Move values from lower triangle to upper triangle
                for i in range(df.shape[0]):
                    for j in range(i):
                        upper = df.iat[j, i]
                        lower = df.iat[i, j]
                        if pd.isna(upper) and not pd.isna(lower):
                            df.iat[j, i] = lower
            
            # Ensure the DataFrame is of float type to accept NaN values
            df = df.astype(float)

            if filename in ["SOLV-fp.xlsx", "INTER.xlsx"]:
                # For SOLV-fp.xlsx and INTER.xlsx, diagonal and below-diagonal must be NaN (upper triangle, k=1)
                df = df.where(np.triu(np.ones(df.shape), k=1).astype(bool))
            else:
                # For all other files, only below-diagonal must be NaN (upper triangle, k=0)
                df = df.where(np.triu(np.ones(df.shape), k=0).astype(bool))
            
            cleaned_sheets[sheet_name] = df

        # Save cleaned sheets
        with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
            for sheet_name, df in cleaned_sheets.items():
                df.to_excel(writer, sheet_name=sheet_name)

    # Step 3: Final relabeling using REF sheet from N-body Summary
    final_labels = extract_final_fragment_labels_from_summary(LEDAW_output_path_nbody)

    if final_labels:
        for filename in files_to_process:
            full_path = os.path.join(LEDAW_output_path_two_body, filename)
            if not os.path.exists(full_path):
                continue

            sheets = pd.read_excel(full_path, sheet_name=None, index_col=0)
            relabeled_sheets = {}

            for sheet_name, df in sheets.items():
                if df.empty or len(df) != len(final_labels):
                    relabeled_sheets[sheet_name] = df
                    continue

                # Ensure index and columns are numeric before assigning new labels
                df.index = final_labels
                df.columns = final_labels

                relabeled_sheets[sheet_name] = df

            with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                for sheet_name, df in relabeled_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name)

    print("Fragment label adjustments and cleanup completed.")


def engine_LED_two_body(one_body_orcaout_filenames, two_body_orcaout_directory, conversion_factor, method, LEDAW_output_path_two_body, reduced_relabel_mapping=None, LEDAW_output_path_nbody=None):
    """Engine function to process LED two-body interaction energy matrices and provide summaries."""

    normalized_LEDAW_output_path_two_body = normalize_path(LEDAW_output_path_two_body)
    normalized_two_body_orcaout_directory = normalize_path(two_body_orcaout_directory)
    normalized_one_body_orcaout_filenames = [normalize_path(f) for f in one_body_orcaout_filenames]
    os.makedirs(normalized_LEDAW_output_path_two_body, exist_ok=True)

    # Detect if BSSE is present in the one-body files
    bsse_found = detect_bsse_from_onebody_files(normalized_one_body_orcaout_filenames)

    if bsse_found:
        # Step 1: Generate remapped consecutive label set (1..N) for INTER, SOLV, ELPREP
        two_body_labels, fragment_index_map, ordered_one_body_files = generate_pairwise_fragment_index_map(normalized_one_body_orcaout_filenames, normalized_two_body_orcaout_directory)

        # Step 2: Find matching (real+ghost) file pairs
        bsse_file_pair_map = generate_bsse_onebody_file_pair(two_body_files=get_two_body_filenames(normalized_two_body_orcaout_directory), one_body_files=normalized_one_body_orcaout_filenames,
            fragment_index_map=fragment_index_map, extract_real_coords_from_twobody_file=extract_real_coords_from_twobody_file, extract_real_and_ghost_coords=extract_real_and_ghost_coords,
            fragments_equal=fragments_equal, tol=1e-3)

        # Step 3: BSSE-corrected solvation contribution and electronic preparation
        solvation_matrix, total_solvation_energy = populate_twobody_total_solvation_matrices_bsse(bsse_file_pair_map=bsse_file_pair_map, two_body_orcaout_directory=normalized_two_body_orcaout_directory,
            conversion_factor=conversion_factor, two_body_labels=two_body_labels, LEDAW_output_path_two_body=normalized_LEDAW_output_path_two_body, method=method)

        nonaggregated_el_prep = populate_twobody_elprep_matrices_bsse(ordered_one_body_files, normalized_two_body_orcaout_directory,
            conversion_factor, method, two_body_labels, bsse_file_pair_map, normalized_LEDAW_output_path_two_body)

    else:
        # Step 4: Label and map fragments and files
        bsse_file_pair_map = None
        ordered_one_body_files = normalized_one_body_orcaout_filenames
        label_mapping = extract_coords_from_one_body_files(ordered_one_body_files)
        two_body_labels = extract_labels_from_two_body_files(normalized_two_body_orcaout_directory, label_mapping)

        # Step 5: BSSE-uncorrected solvation contribution and electronic preparation
        solvation_matrix, total_solvation_energy = populate_twobody_total_solvation_matrices_non_bsse(ordered_one_body_files, normalized_two_body_orcaout_directory,
            conversion_factor, two_body_labels, normalized_LEDAW_output_path_two_body, method)

        nonaggregated_el_prep = populate_twobody_elprep_matrices_non_bsse(ordered_one_body_files, normalized_two_body_orcaout_directory,
            conversion_factor, method, two_body_labels, normalized_LEDAW_output_path_two_body)

    # Step 6: INTER matrices
    populate_twobody_inter_matrices(ordered_one_body_files, normalized_two_body_orcaout_directory, conversion_factor,
        two_body_labels=two_body_labels, LEDAW_output_path_two_body=normalized_LEDAW_output_path_two_body)

    # Step 7: Summary matrices
    calculate_twobody_standard_LED_summary_matrices(normalized_LEDAW_output_path_two_body, method)

    # Step 8: Solvation Contribution STD
    compute_and_write_solv_std(normalized_LEDAW_output_path_two_body, method, nonaggregated_el_prep)

    # Step 9: fp-LED matrices
    calculate_twobody_fpLED_matrices(normalized_LEDAW_output_path_two_body, method)

    # Step 10: Cleanup
    cleanup_and_reindex_all_excels(normalized_LEDAW_output_path_two_body, reduced_relabel_mapping=reduced_relabel_mapping, LEDAW_output_path_nbody=LEDAW_output_path_nbody)

    print('\n')
    print('*'*125)
    print(f"  Two-body LED analyses were terminated NORMALLY. Summary matrices are saved in '{normalized_LEDAW_output_path_two_body}'")
    print('*'*125)